# -*- coding: utf-8 -*-
"""
=====================================================================
 ANÁLISE DE MÉTRICAS TREBLE — relatório de período (equipe completa)
 Nexforce | RevOps
=====================================================================

O QUE ESTA VERSÃO ENTREGA
-------------------------
Base de verdade = aba "atendimento-treble" (dados por conversa com vendedor
real, timestamps de atribuição e 1ª resposta, quem mandou a última mensagem
e tipo de encerramento). Tudo é validado contra ela.

Análises (relatório semanal — layout Utilização Geral Treble):
  01. Conversas: respondidas x sem resposta do vendedor (base atendimento).
  02. Chegada ao vendedor x conversas atendidas — barras absolutas sobrepostas + empilhado %.
  03. Conversas recebidas por vendedor + % de conversas respondidas por vendedor.
  04. Tempo mediano de 1ª resposta (h).
  05. Padrão de resíduo dos contatos atribuídos (visão geral):
      composição por padrão + respondidos x sem resposta por padrão.
  06. Preenchimento dos campos de qualificação, separado por caminho:
      "Sou novo por aqui" e "Já sou cliente".

  Correção de consistência:
  - Base canônica: apenas conversas com agente reconhecido (exclui as contas em
    EXCLUIR_AGENTES_NORM e linhas sem agente). Todos os contadores usam o mesmo
    denominador — sem mais divergência entre totais de análises diferentes.

  Observações de cálculo:
  - CPF/CNPJ: um contato conta como preenchido se tiver QUALQUER um dos dois.
  - Número de versões do fluxo é variável e detectado automaticamente.
  - Blocos auxiliares (sessões, retorno, inbox, transferências, território)
    continuam sendo calculados e exportados em abas, mas NÃO entram no painel
    do relatório.

Saídas:
  - analise_treble_resultado.xlsx
      -> 1ª aba "00_PAINEL" reúne TODOS os gráficos em ordem narrativa
  - pasta graficos/equipe/ com os PNGs individuais

REGRAS IMPORTANTES
------------------
* Telefone = identificador; match por "núcleo nacional" (ignora o 55), porque
  cada aba grava em um formato (geral sem 55 / versão com 55 / atendimento "+55 ").
* Números de teste (TELEFONES_TESTE) e contas internas (EXCLUIR_AGENTES_NORM)
  são EXCLUÍDOS de todas as análises.
* Nomes de vendedor vêm inconsistentes (maiúsc./minúsc./acento) e são
  canonizados para a lista oficial.
"""

import os
import re
import unicodedata
import pandas as pd
import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# Funcoes de metodologia (secoes 05/06) centralizadas no modulo compartilhado:
import sys as _sys, os as _os
_sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
from _analise_comum import (
    _CATS_RESIDUO, _classificar_padrao_residuo,
    padrao_residuo_atribuidos, analise_caminhos,
)


# =====================================================================
# CONFIGURAÇÃO
# =====================================================================

ARQUIVO = None   # definido pelo CLI (período). Ex.: "analises/[período] (geral)/Dados_Treble_Semana.xlsx"
# Data de referência da execução (a segunda-feira em que roda). None = usa hoje.
# A análise sempre cobre a SEMANA ANTERIOR completa (domingo a sábado).
DATA_REFERENCIA = None      # ex.: "2026-06-15" (segunda de execução)
# Janela explícita da semana analisada (domingo a sábado). Tem PRIORIDADE.
# Deixe ambos None para detectar automaticamente a partir dos dados; se também
# não houver dados datados, cai na semana anterior a DATA_REFERENCIA/hoje.
#
# >>> EXECUÇÃO AUTOMÁTICA SEMANAL: mantenha os dois como None.
# >>> REPROCESSAR uma semana antiga: preencha com as datas daquela semana.
SEMANA_INICIO = None   # AAAA-MM-DD ou None (definido pelo CLI a partir do período)
SEMANA_FIM    = None   # AAAA-MM-DD ou None

# NÚMERO DE VERSÕES DO FLUXO É VARIÁVEL (1, 2, ... N). Não é preciso configurar
# nada: carregar_arquivo() detecta toda aba com a coluna "Celular" como uma
# versão, e consolidar_versoes() concatena TODAS automaticamente. Uma semana
# pode ter 1 versão ou 30; o código se adapta sozinho.
ARQUIVO_SAIDA = "relatorio_equipe_periodo.xlsx"   # gravado na raiz do projeto
PASTA_GRAFICOS = "graficos/equipe"

# --- Exclusões ---
EXCLUIR_TELEFONES_TESTE = True
TELEFONES_TESTE = [
    # Telefones de teste/QA a excluir de TODAS as análises. Formato "+55DDDNÚMERO".
    # Preencha com os seus; deixe a lista vazia se não houver.
]
# Contas internas/bots a desconsiderar (normalizadas: sem acento/espaço/maiúsc.)
EXCLUIR_AGENTES_NORM = set()  # ex.: {"contadeteste", "usuariosistema"}

# Lista oficial de vendedores — SUBSTITUA pelos nomes reais do seu fluxo
# (exatamente como aparecem no campo 'agent' da base de atendimento).
# Os nomes abaixo são FICTÍCIOS, apenas de exemplo.
AGENTES_CANONICOS = [
    "Ana Souza", "Bruno Costa", "Carla Mendes", "Daniel Rocha",
    "Eduarda Nunes", "Felipe Ramos", "Beatriz Dias", "Henrique Alves",
    "Isabela Pinto", "Larissa Gomes", "Marcos Vieira", "Renata Oliveira",
]

# Vendedores desligados: mantenha-os em AGENTES_CANONICOS (para não descartar o
# histórico deles), e liste aqui para tirá-los dos gráficos por vendedor nas
# janelas em que não receberam nenhuma conversa.
AGENTES_DESLIGADOS = set()  # ex.: {"Bruno Costa"}

JANELA_24H = 24.0
JANELA_72H = 72.0
LIMIAR_ESTAGNADA_H = 70.0
MSG_HANDOVER = "Agradeço as informações"
MSG_FORA_HORARIO = "No momento, nosso time não está disponível"

# Paleta
COR_AZUL = "#1F5C99"; COR_NAVY = "#0F2A47"; COR_CIANO = "#5FA8DC"
COR_CINZA = "#AAB4BF"; COR_ALERTA = "#C0392B"; COR_OK = "#2E8B57"
COR_AMBAR = "#E08E0B"


# =====================================================================
# UTILITÁRIOS DE BASE
# =====================================================================

def corrigir_mojibake(s):
    if isinstance(s, str) and ("Ã" in s or "Â" in s):
        try:
            return s.encode("latin1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            return s
    return s


def _norm(txt):
    txt = corrigir_mojibake(str(txt))
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", txt.lower())


def achar_coluna(df, *candidatos):
    mapa = {_norm(c): c for c in df.columns}
    for cand in candidatos:
        chave = _norm(cand)
        if chave in mapa:
            return mapa[chave]
        for k, real in mapa.items():
            if chave in k or k in chave:
                return real
    return None


def normalizar_telefone(serie):
    return (serie.astype(str).str.replace(r"\.0$", "", regex=True)
            .str.replace(r"\D", "", regex=True)
            .replace({"": np.nan, "nan": np.nan}))


def nucleo_telefone(num):
    d = re.sub(r"\D", "", str(num))
    if len(d) >= 12 and d.startswith("55"):
        d = d[2:]
    return d


_NUCLEOS_TESTE = {nucleo_telefone(t) for t in TELEFONES_TESTE}


def _mascara_nao_teste(serie_tel):
    return ~serie_tel.map(nucleo_telefone).isin(_NUCLEOS_TESTE)


def _corrigir_df(df):
    df = df.copy()
    df.columns = [corrigir_mojibake(c) for c in df.columns]
    for c in df.columns:
        if df[c].dtype == object or str(df[c].dtype) == "str":
            df[c] = df[c].map(corrigir_mojibake)
    return df


# --- canonização de nomes de vendedor ---
_MAPA_CANON = {_norm(a): a for a in AGENTES_CANONICOS}


def canonizar_agente(nome):
    """Mapeia variações ('ana souza', 'BRUNO COSTA') para o nome oficial.
    Retorna None se for conta a excluir ou vazio."""
    if pd.isna(nome):
        return None
    n = _norm(nome)
    if not n or n in EXCLUIR_AGENTES_NORM:
        return None
    if n in _MAPA_CANON:
        return _MAPA_CANON[n]
    # casamento por tokens (sobrenome/primeiro nome)
    tokens = {t for t in re.split(r"[^a-z0-9]+", _norm_spaces(nome)) if len(t) >= 4}
    for canon in AGENTES_CANONICOS:
        ctok = {t for t in re.split(r"[^a-z0-9]+", _norm_spaces(canon)) if len(t) >= 4}
        if tokens & ctok:
            return canon
    return str(nome).strip().title()


def _norm_spaces(txt):
    txt = corrigir_mojibake(str(txt))
    txt = unicodedata.normalize("NFKD", txt).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", txt.lower().strip())


def agente_de_email(email):
    """Deriva o vendedor canônico a partir do e-mail (ex.: ana.souza@... -> Ana Souza)."""
    if pd.isna(email) or "@" not in str(email):
        return None
    local = str(email).split("@")[0]
    local = re.sub(r"\d+", "", local)
    tokens = {t for t in re.split(r"[^a-z0-9]+", _norm_spaces(local)) if len(t) >= 4}
    melhor = None
    for canon in AGENTES_CANONICOS:
        ctok = {t for t in re.split(r"[^a-z0-9]+", _norm_spaces(canon)) if len(t) >= 4}
        if tokens & ctok:
            melhor = canon
            break
    if melhor and _norm(melhor) in EXCLUIR_AGENTES_NORM:
        return None
    return melhor


# =====================================================================
# CARGA E CLASSIFICAÇÃO DE ABAS (por colunas, não por nome)
# =====================================================================

def carregar_arquivo(caminho=ARQUIVO):
    xls = pd.ExcelFile(caminho)
    out = {"gerais": {}, "versoes": {}, "atendimento": None, "inbox": None}
    for aba in xls.sheet_names:
        df = _corrigir_df(pd.read_excel(xls, sheet_name=aba))
        cols = {_norm(c) for c in df.columns}
        if {"agent", "lastmessagesender"} & cols and "finishtype" in cols:
            out["atendimento"] = df
        elif achar_coluna(df, "Total de bate-papos recebidos") or \
                ({"encontro", "contatosdistintos"} <= cols):
            out["inbox"] = df
        elif achar_coluna(df, "user_cellphone") is not None:
            out["gerais"][aba] = df
        elif achar_coluna(df, "Celular") is not None:
            m = re.search(r"(\d+)", str(aba))
            out["versoes"][m.group(1) if m else aba] = df
        else:
            print(f"[aviso] aba '{aba}' não reconhecida.")
    print(f"Gerais: {list(out['gerais'].keys())} | Versões: {list(out['versoes'].keys())} "
          f"| atendimento: {'OK' if out['atendimento'] is not None else '-'} "
          f"| inbox: {'OK' if out['inbox'] is not None else '-'}")
    return out


def consolidar_gerais(dados, abas=None, excluir_teste=EXCLUIR_TELEFONES_TESTE):
    fonte = dados["gerais"]
    if abas:
        fonte = {k: v for k, v in fonte.items() if k in abas}
    frames = []
    for nome, df in fonte.items():
        d = df.copy(); d["origem_aba"] = nome
        d["telefone"] = normalizar_telefone(d[achar_coluna(d, "user_cellphone")])
        frames.append(d)
    if not frames:
        return pd.DataFrame()
    g = pd.concat(frames, ignore_index=True)
    for c in ["session_started_timestamp", "session_finished_timestamp",
              "first_message_timestamp", "last_message_timestamp"]:
        real = achar_coluna(g, c)
        if real:
            g[real] = pd.to_datetime(g[real], errors="coerce")
    if excluir_teste:
        antes = len(g); g = g[_mascara_nao_teste(g["telefone"])].copy()
        print(f"[gerais] testes removidos: {antes - len(g)} (restaram {len(g)})")
    return g


def consolidar_versoes(dados, versoes=None, excluir_teste=EXCLUIR_TELEFONES_TESTE):
    fonte = dados["versoes"]
    if versoes:
        fonte = {k: v for k, v in fonte.items() if k in versoes}
    frames = []
    for versao, df in fonte.items():
        d = df.copy(); d["versao"] = versao
        d["telefone"] = normalizar_telefone(d[achar_coluna(d, "Celular")])
        col_ult = achar_coluna(d, "ultima atividade")
        if col_ult:
            d["ultima_atividade_dt"] = pd.to_datetime(d[col_ult], errors="coerce")
        frames.append(d)
    if not frames:
        return pd.DataFrame()
    v = pd.concat(frames, ignore_index=True)
    if excluir_teste:
        antes = len(v); v = v[_mascara_nao_teste(v["telefone"])].copy()
        print(f"[versões] testes removidos: {antes - len(v)} (restaram {len(v)})")
    return v


def consolidar_atendimento(dados, excluir_teste=EXCLUIR_TELEFONES_TESTE):
    """Prepara a base mestre de atendimento, com campos derivados."""
    df = dados.get("atendimento")
    if df is None:
        return pd.DataFrame()
    d = df.copy()
    d["telefone"] = normalizar_telefone(d[achar_coluna(d, "phone")])
    d["nucleo"] = d["telefone"].map(nucleo_telefone)
    # datas
    for c in ["created_at", "assigned_at", "finished_at", "last_transfer_time",
              "agent_first_message", "last_message"]:
        real = achar_coluna(d, c)
        if real:
            d[real] = pd.to_datetime(d[real], errors="coerce")
    col_assign = achar_coluna(d, "assigned_at")
    col_first = achar_coluna(d, "agent_first_message")
    col_ltt = achar_coluna(d, "last_transfer_time")
    col_ltf = achar_coluna(d, "last_transfer_from")
    col_sender = achar_coluna(d, "last_message_sender")
    col_finish = achar_coluna(d, "finish_type")

    # vendedor canônico (1º atribuído) e receptor de transferência
    d["agente"] = d[achar_coluna(d, "agent")].map(canonizar_agente)
    # Dicionário (06/07/2026): last_transfer_from = quem ORIGINOU a transferência
    d["origem_ultima_transfer"] = d[col_ltf].map(agente_de_email) if col_ltf else None

    # tempo de 1ª resposta e flags
    d["tempo_1a_resposta_h"] = (d[col_first] - d[col_assign]).dt.total_seconds() / 3600
    d["respondido"] = d[col_first].notna()
    d["respondido_24h"] = d["respondido"] & (d["tempo_1a_resposta_h"] <= JANELA_24H)
    d["transferido"] = d[col_ltt].notna() if col_ltt else False
    d["sender"] = d[col_sender].astype(str).str.upper().str.strip() if col_sender else None
    d["finish"] = d[col_finish].astype(str).str.upper().str.strip() if col_finish else None

    # exclui contas internas (agente None que veio de exclusão) e telefones de teste
    if excluir_teste:
        antes = len(d); d = d[_mascara_nao_teste(d["telefone"])].copy()
        print(f"[atendimento] testes removidos por telefone: {antes - len(d)}")
    # exclui linhas cujo agente foi excluído (contas em EXCLUIR_AGENTES_NORM)
    # ou não tem agente registrado — garante denominador único em todas as análises.
    antes2 = len(d); d = d[d["agente"].notna()].copy()
    print(f"[atendimento] excluídos sem agente canônico: {antes2 - len(d)} (restaram {len(d)})")
    return d


def consolidar_inbox(dados):
    df = dados.get("inbox")
    if df is None:
        return pd.DataFrame()
    d = df.copy()
    col_data = achar_coluna(d, "Encontro", "data", "date")
    if col_data:
        d[col_data] = pd.to_datetime(d[col_data], errors="coerce")
        d = d.rename(columns={col_data: "data"})
    return d


# =====================================================================
# A. ATENDIMENTO — GERAL
# =====================================================================

def atendimento_geral(at):
    total = len(at)
    com_agente = at["agente"].notna().sum()
    respond = int(at["respondido"].sum())
    resp24 = int(at["respondido_24h"].sum())
    nao_resp = total - respond
    auto = int((at["finish"] == "AUTO").sum()) if "finish" in at else 0
    manual = int((at["finish"] == "MANUAL").sum()) if "finish" in at else 0
    transf = int(at["transferido"].sum())
    tmed = round(at.loc[at["respondido"], "tempo_1a_resposta_h"].median(), 2)
    resumo = pd.DataFrame([
        ("Conversas atendidas (total)", total),
        ("Conversas respondidas por vendedor", respond),
        ("Conversas SEM resposta do vendedor", nao_resp),
        ("% sem resposta", round(100 * nao_resp / total, 1) if total else 0),
        ("Respondidas dentro de 24h", resp24),
        ("% respondidas em 24h (do total)", round(100 * resp24 / total, 1) if total else 0),
        ("% respondidas em 24h (das respondidas)", round(100 * resp24 / respond, 1) if respond else 0),
        ("Tempo mediano de 1ª resposta (h)", tmed),
        ("Encerradas em AUTO", auto),
        ("Encerradas em MANUAL", manual),
        ("% encerramento AUTO", round(100 * auto / total, 1) if total else 0),
        ("Conversas transferidas", transf),
    ], columns=["Métrica", "Valor"])
    return resumo


def panorama_time(at):
    """Tabela de panorama geral do time: indicadores consolidados de todos os
    vendedores com agente atribuído. Exibida no início do relatório."""
    base = at[at["agente"].notna()].copy()
    n_vend = base["agente"].nunique()
    total  = len(base)
    if total == 0:
        return pd.DataFrame()
    respond   = int(base["respondido"].sum())
    nao_resp  = total - respond
    resp24    = int(base["respondido_24h"].sum())
    tmed      = base.loc[base["respondido"], "tempo_1a_resposta_h"].median()
    transf_out = int(base["transferido"].sum())
    transf_in  = int(base["origem_ultima_transfer"].notna().sum()) if "origem_ultima_transfer" in base else 0
    med_chats  = round(total / n_vend, 1) if n_vend else 0
    rows = [
        ("Vendedores ativos na semana",  n_vend),
        ("Total de conversas atribuídas", total),
        ("Média de conversas por vendedor", med_chats),
        ("Conversas respondidas",         respond),
        ("Conversas sem resposta",         nao_resp),
        ("% sem resposta",                round(100 * nao_resp / total, 1)),
        ("Respondidas dentro de 24h",     resp24),
        ("% respondidas em 24h",          round(100 * resp24 / total, 1)),
        ("Tempo mediano de 1ª resposta (h)", round(tmed, 2) if pd.notna(tmed) else "—"),
        ("Conversas transferidas (out)",  transf_out),
        ("Recebidas via transferência",   transf_in),
    ]
    return pd.DataFrame(rows, columns=["Métrica", "Valor"])


def atendimento_last_sender(at):
    if "sender" not in at:
        return pd.DataFrame()
    t = (at["sender"].value_counts(dropna=False).rename_axis("last_message_sender")
         .reset_index(name="conversas"))
    t["%"] = round(100 * t["conversas"] / t["conversas"].sum(), 1)
    return t


def atendimento_finish_type(at):
    if "finish" not in at:
        return pd.DataFrame()
    t = (at["finish"].value_counts(dropna=False).rename_axis("finish_type")
         .reset_index(name="conversas"))
    t["%"] = round(100 * t["conversas"] / t["conversas"].sum(), 1)
    return t


# =====================================================================
# B. RESPOSTA NA JANELA 24h — POR DIA
# =====================================================================

def resposta_24h_diaria(at):
    col_assign = achar_coluna(at, "assigned_at")
    base = at.dropna(subset=[col_assign]).copy()
    base["dia"] = base[col_assign].dt.floor("D")
    agg = base.groupby("dia").agg(
        chegaram=("agente", "size"),
        respondidos_24h=("respondido_24h", "sum"),
    )
    agg["%_em_24h"] = round(100 * agg["respondidos_24h"] / agg["chegaram"], 1)
    return agg.reset_index()


def resposta_24h_semanal(at):
    """Agrupa chegadas x atendimentos em 24h por semana dom-sáb (âncora W-SAT)."""
    col_assign = achar_coluna(at, "assigned_at")
    base = at.dropna(subset=[col_assign]).copy()
    base["semana_ini"] = base[col_assign].dt.to_period("W-SAT").dt.start_time
    agg = base.groupby("semana_ini").agg(
        chegaram=("agente", "size"),
        respondidos_24h=("respondido_24h", "sum"),
    )
    agg["%_em_24h"] = round(100 * agg["respondidos_24h"] / agg["chegaram"], 1)
    return agg.reset_index().rename(columns={"semana_ini": "dia"})


# =====================================================================
# C. RETORNO DE CLIENTES (números repetidos)
# =====================================================================

def retorno_clientes(at):
    col_created = achar_coluna(at, "created_at")
    g = at.dropna(subset=["nucleo"]).groupby("nucleo")
    cont = g.agg(contatos=("nucleo", "size"),
                 primeiro=(col_created, "min"),
                 ultimo=(col_created, "max")).reset_index()
    cont["dias_entre_1o_e_ultimo"] = (cont["ultimo"] - cont["primeiro"]).dt.total_seconds() / 86400
    cont["dias_entre_1o_e_ultimo"] = cont["dias_entre_1o_e_ultimo"].round(1)

    unicos = len(cont)
    voltaram = int((cont["contatos"] >= 2).sum())
    resumo = pd.DataFrame([
        ("Contatos únicos (telefones)", unicos),
        ("Conversas totais", int(cont["contatos"].sum())),
        ("Contatos que voltaram (>=2 conversas)", voltaram),
        ("% de retorno", round(100 * voltaram / unicos, 1) if unicos else 0),
        ("Máx. de conversas de um mesmo contato", int(cont["contatos"].max()) if unicos else 0),
    ], columns=["Métrica", "Valor"])

    dist = (cont["contatos"].clip(upper=4)
            .map({1: "1 conversa", 2: "2 conversas", 3: "3 conversas", 4: "4+ conversas"})
            .value_counts().rename_axis("faixa").reset_index(name="contatos"))
    ordem = ["1 conversa", "2 conversas", "3 conversas", "4+ conversas"]
    dist["ord"] = dist["faixa"].map({k: i for i, k in enumerate(ordem)})
    dist = dist.sort_values("ord").drop(columns="ord")

    recorrentes = (cont[cont["contatos"] >= 2]
                   .sort_values("contatos", ascending=False)
                   .rename(columns={"nucleo": "telefone_nucleo"}))
    return {"resumo_retorno": resumo, "distribuicao": dist, "recorrentes": recorrentes}


# =====================================================================
# D. POR VENDEDOR  (núcleo do pedido do chefe)
# =====================================================================

def por_vendedor(at):
    # quem ORIGINOU transferências (last_transfer_from) → conta como transferiu_out
    transfer_out = (at["origem_ultima_transfer"].dropna().value_counts()
                    if "origem_ultima_transfer" in at else pd.Series(dtype=int))

    linhas = []
    base = at[at["agente"].notna()]
    for ag in AGENTES_CANONICOS:
        sub = base[base["agente"] == ag]
        n = len(sub)
        if n == 0 and transfer_out.get(ag, 0) == 0:
            continue
        respond = int(sub["respondido"].sum())
        resp24 = int(sub["respondido_24h"].sum())
        nao = n - respond
        # transferiu out: foi 1º agente, houve transferência e o receptor final != ele
        # recebeu via transfer: conversa dele com transferência originada por OUTRO vendedor
        recebeu_transfer = int((sub["transferido"] & (sub["origem_ultima_transfer"] != ag)).sum())
        auto = int((sub["finish"] == "AUTO").sum())
        manual = int((sub["finish"] == "MANUAL").sum())
        ia = int((sub["sender"] == "IA").sum())
        agent_s = int((sub["sender"] == "AGENT").sum())
        user_s = int((sub["sender"] == "USER").sum())
        tmed = round(sub.loc[sub["respondido"], "tempo_1a_resposta_h"].median(), 2) if respond else np.nan
        linhas.append({
            "vendedor": ag,
            "chats_recebidos": n,
            "respondidos": respond,
            "nao_respondidos": nao,
            "%_nao_respondidos": round(100 * nao / n, 1) if n else 0,
            "respondidos_24h": resp24,
            "%_em_24h": round(100 * resp24 / n, 1) if n else 0,
            "tempo_med_1a_resp_h": tmed,
            "transferiu_out": int(transfer_out.get(ag, 0)),
            "recebidos_via_transfer": recebeu_transfer,
            "fim_AUTO": auto,
            "fim_MANUAL": manual,
            "ult_msg_IA": ia,
            "ult_msg_AGENT": agent_s,
            "ult_msg_USER": user_s,
        })
    tab = pd.DataFrame(linhas).sort_values("chats_recebidos", ascending=False).reset_index(drop=True)
    return tab


def tempo_resposta_semanal_time(at):
    """Tempo mediano de 1ª resposta (h) do time por semana — linha única do gráfico de evolução.

    A mediana é calculada sobre TODAS as respostas da semana (não é média das
    medianas por vendedor). Análise não segmentada por decisão de 2026-07-03.
    """
    col_assign = achar_coluna(at, "assigned_at")
    if not col_assign:
        return pd.DataFrame()
    base = at[at["respondido"] & at["agente"].notna()].dropna(subset=[col_assign]).copy()
    if base.empty:
        return pd.DataFrame()
    base["semana_ini"] = base[col_assign].dt.to_period("W-SAT").dt.start_time
    return (base.groupby("semana_ini")["tempo_1a_resposta_h"]
            .median().round(2).reset_index()
            .rename(columns={"tempo_1a_resposta_h": "tempo_med_h",
                             "semana_ini": "dia"}))


def transferencias(at):
    """Resumo de transferências: out (quem originou, last_transfer_from) x in (agente final que recebeu)."""
    base = at[at["agente"].notna()]
    inn = (base[base["transferido"] & (base["origem_ultima_transfer"] != base["agente"])]
           .groupby("agente").size().rename("recebidos_via_transfer"))
    out = (at["origem_ultima_transfer"].dropna().value_counts().rename("transferiu_out"))
    tab = pd.concat([out, inn], axis=1).fillna(0).astype(int)
    tab = tab.reindex([a for a in AGENTES_CANONICOS if a in tab.index]).dropna(how="all")
    return tab.reset_index(names="vendedor")


# =====================================================================
# PADRÃO DE RESÍDUO DOS CONTATOS ATRIBUÍDOS
# Cruza a base de atendimento (quem chegou a um vendedor) com a classe de
# resíduo informada no fluxo (abas de versão), pelo núcleo do telefone.
# Visão GERAL (não aponta para nenhum vendedor específico).
# =====================================================================





# =====================================================================
# D4. FICHA INDIVIDUAL POR VENDEDOR (uma aba por vendedor)
# =====================================================================

def _linhas_ficha(sub, others, n_out_ativos, ti_v, ti_others_avg):
    """Monta as linhas (Métrica, valor do vendedor, média do time)."""
    def pct(num, den):
        return round(100 * num / den, 1) if den else np.nan
    ns, no = len(sub), len(others)
    rs = int(sub["respondido"].sum()) if ns else 0
    ro = int(others["respondido"].sum()) if no else 0
    # semântica corrigida (06/07/2026): origem != agente → o vendedor RECEBEU via transfer
    transf_v = int((sub["transferido"] & (sub["origem_ultima_transfer"] != sub["agente"])).sum()) if ns else 0
    transf_o = int((others["transferido"] & (others["origem_ultima_transfer"] != others["agente"])).sum()) if no else 0
    tmed_v = round(sub.loc[sub["respondido"], "tempo_1a_resposta_h"].median(), 2) if rs else np.nan
    tmed_o = round(others.loc[others["respondido"], "tempo_1a_resposta_h"].median(), 2) if ro else np.nan
    media = (lambda x: round(x / n_out_ativos, 1) if n_out_ativos else 0)
    return [
        ("Conversas recebidas", ns, media(no)),
        ("Conversas transferidas (out)", transf_v, media(transf_o)),
        ("Recebidas via transferência", ti_v, round(ti_others_avg, 1)),
        ("Conversas respondidas", rs, media(ro)),
        ("% não respondidas", pct(ns - rs, ns), pct(no - ro, no)),
        ("% de conversas atendidas", pct(int(sub["respondido_24h"].sum()), ns), pct(int(others["respondido_24h"].sum()), no)),
        ("Tempo mediano 1ª resposta (h)", tmed_v, tmed_o),
        ("% encerramento AUTO", pct(int((sub["finish"] == "AUTO").sum()), ns), pct(int((others["finish"] == "AUTO").sum()), no)),
        ("% última msg do cliente (sem retorno)", pct(int((sub["sender"] == "USER").sum()), ns), pct(int((others["sender"] == "USER").sum()), no)),
    ]


def fichas_vendedores(at):
    """Retorna {vendedor: DataFrame} — uma ficha por vendedor, comparando com a
    média do time (demais vendedores agrupados)."""
    base = at[at["agente"].isin(AGENTES_CANONICOS)]
    transfer_in = at["origem_ultima_transfer"].dropna().value_counts()  # aqui: quem ORIGINOU
    fichas = {}
    for ag in AGENTES_CANONICOS:
        sub = base[base["agente"] == ag]
        others = base[base["agente"] != ag]
        ti_v = int(transfer_in.get(ag, 0))
        if len(sub) == 0 and ti_v == 0:
            continue  # vendedor sem nenhuma atividade nos dados
        n_out_ativos = others["agente"].nunique()
        ti_others_avg = (sum(int(transfer_in.get(a, 0)) for a in AGENTES_CANONICOS if a != ag) / n_out_ativos) \
            if n_out_ativos else 0
        rows = _linhas_ficha(sub, others, n_out_ativos, ti_v, ti_others_avg)
        df = pd.DataFrame(rows, columns=["Métrica", ag, "Média do time"])
        df["Δ vs time"] = (pd.to_numeric(df[ag], errors="coerce")
                           - pd.to_numeric(df["Média do time"], errors="coerce")).round(1)
        fichas[ag] = df
    return fichas


def grafico_ficha(vendedor, df):
    """Gráfico de barras pareadas (vendedor x média do time) com as métricas em %."""
    labels_full = ["% não respondidos", "% respondidos em 24h",
                   "% encerramento AUTO", "% última msg do cliente (sem retorno)"]
    labels_short = ["% não\nresp.", "% em\n24h", "% AUTO", "% últ.\ncliente"]
    m = {met: (vv, tt) for met, vv, tt in zip(df["Métrica"], df[vendedor], df["Média do time"])}
    v, t, lab = [], [], []
    for lf, ls in zip(labels_full, labels_short):
        if lf in m and not pd.isna(m[lf][0]):
            v.append(m[lf][0]); t.append(m[lf][1]); lab.append(ls)
    if not v:
        return None
    fig, ax = plt.subplots(figsize=(7.5, 4.4))
    x = np.arange(len(lab)); w = 0.38
    b1 = ax.bar(x - w/2, v, w, color=COR_AZUL, label=vendedor)
    b2 = ax.bar(x + w/2, t, w, color=COR_CINZA, label="Média do time")
    ax.set_xticks(x); ax.set_xticklabels(lab, fontsize=9)
    ax.set_ylim(0, 105); ax.grid(axis="y", color="#E6EBF0")
    _ax(ax, f"{vendedor} — desempenho x média do time")
    ax.legend(fontsize=8)
    for b in (b1, b2):
        try:
            ax.bar_label(b, fmt="%g", padding=2, fontsize=8)
        except Exception:
            pass
    return _save(fig, f"V_{_norm(vendedor)}.png")


# =====================================================================
# E. INBOX DIÁRIO
# =====================================================================

def inbox_diario(ib):
    if not len(ib):
        return pd.DataFrame()
    col_tot = achar_coluna(ib, "Total de bate-papos recebidos")
    col_dist = achar_coluna(ib, "Contatos distintos")
    col_tr = achar_coluna(ib, "Conversas transferidas")
    cols = {"data": "data"}
    if col_tot: cols[col_tot] = "total_recebidos"
    if col_dist: cols[col_dist] = "contatos_distintos"
    if col_tr: cols[col_tr] = "transferidas"
    out = ib.rename(columns=cols)
    manter = [c for c in ["data", "total_recebidos", "contatos_distintos", "transferidas"] if c in out]
    return out[manter].sort_values("data")


# =====================================================================
# ===== ANÁLISES ANTERIORES (mantidas) — log de sessões + versões =====
# =====================================================================

def visao_geral_periodo(g):
    col_status = achar_coluna(g, "session_status")
    col_ver = achar_coluna(g, "conversation_version")
    col_lastmsg = achar_coluna(g, "last_message")
    total = len(g); unicos = g["telefone"].nunique()
    sv = g[col_status].value_counts(dropna=False) if col_status else pd.Series(dtype=int)
    nh = int(sv.get("HumanHandover", 0)); na = int(sv.get("AI", 0))
    fora = int(g[col_lastmsg].astype(str).str.contains(MSG_FORA_HORARIO, na=False).sum()) if col_lastmsg else 0
    resumo = pd.DataFrame([
        ("Sessões no período", total), ("Contatos únicos", unicos),
        ("Repassadas a vendedor (HumanHandover)", nh),
        ("Resolvidas/abandonadas no bot (AI)", na),
        ("Taxa de repasse (%)", round(100 * nh / total, 1) if total else 0),
        ("Sessões fora de horário", fora),
    ], columns=["Métrica", "Valor"])
    vd = pd.DataFrame()
    if col_ver:
        vd = g[col_ver].value_counts(dropna=False).rename_axis("conversation_version").reset_index(name="sessoes")
    return {"resumo": resumo, "distribuicao_versao": vd, "_status": {"AI": na, "HumanHandover": nh}}


def funil_por_no(g):
    col_no = achar_coluna(g, "last_node_id"); col_status = achar_coluna(g, "session_status")
    if not col_no:
        return pd.DataFrame()
    tab = g.groupby([col_no, col_status]).size().unstack(fill_value=0)
    tab["total"] = tab.sum(axis=1)
    tab = tab.sort_values("total", ascending=False).reset_index()
    tab["%_do_total"] = round(100 * tab["total"] / tab["total"].sum(), 1)
    return tab


def tendencia_temporal(g, freq="D"):
    col_ini = achar_coluna(g, "session_started_timestamp"); col_status = achar_coluna(g, "session_status")
    if not col_ini:
        return pd.DataFrame()
    d = g.dropna(subset=[col_ini]).copy()
    d["periodo"] = d[col_ini].dt.to_period(freq).dt.start_time
    agg = d.groupby("periodo").agg(sessoes=("telefone", "size"), contatos_unicos=("telefone", "nunique"))
    if col_status:
        rep = d[d[col_status] == "HumanHandover"].groupby("periodo").size()
        agg["repasses"] = rep.reindex(agg.index).fillna(0).astype(int)
        agg["taxa_repasse_%"] = round(100 * agg["repasses"] / agg["sessoes"], 1)
    return agg.reset_index()




def qualidade_dados(v):
    achados = []
    col_res = achar_coluna(v, "hubspot_tipo_de_residuo")
    if col_res:
        validos = {"industrial", "hospitalar"}
        inval = v[col_res].dropna()
        inval = inval[~inval.astype(str).str.strip().str.lower().isin(validos)]
        achados.append(("tipo_de_residuo inesperado", len(inval), ", ".join(map(str, inval.unique()[:10]))))
    col_cpf = achar_coluna(v, "hubspot_cpf"); col_cnpj = achar_coluna(v, "hubspot_cnpj")
    if col_cpf and col_cnpj:
        achados.append(("CPF e CNPJ juntos", int((v[col_cpf].notna() & v[col_cnpj].notna()).sum()), ""))
    flags = pd.DataFrame(achados, columns=["achado", "qtd", "exemplos"])
    col_uf = achar_coluna(v, "hubspot_state")
    est = pd.DataFrame()
    if col_uf:
        est = v[col_uf].value_counts(dropna=False).rename_axis("estado_original").reset_index(name="qtd")
    return {"flags_qualidade": flags, "estado_valores_originais": est}


_UF_MAP = {"saopaulo": "SP", "sp": "SP", "riodejaneiro": "RJ", "rj": "RJ", "minasgerais": "MG", "mg": "MG",
           "parana": "PR", "pr": "PR", "santacatarina": "SC", "sc": "SC", "riograndedosul": "RS", "rs": "RS",
           "bahia": "BA", "ba": "BA", "espiritosanto": "ES", "es": "ES", "goias": "GO", "go": "GO",
           "distritofederal": "DF", "df": "DF"}


def analise_territorial(v):
    col_uf = achar_coluna(v, "hubspot_state"); col_cid = achar_coluna(v, "hubspot_city")
    out = {}
    if col_uf:
        uf = v[col_uf].map(lambda x: np.nan if pd.isna(x) else _UF_MAP.get(_norm(x), str(x).strip().upper()))
        out["por_estado"] = uf.value_counts(dropna=False).rename_axis("UF").reset_index(name="conversas")
    if col_cid:
        cid = v[col_cid].astype(str).str.strip().str.title()
        out["por_cidade"] = cid.value_counts(dropna=False).rename_axis("cidade").reset_index(name="conversas").head(40)
    return out


def analise_72h(g):
    col_status = achar_coluna(g, "session_status"); col_ini = achar_coluna(g, "session_started_timestamp")
    col_fim = achar_coluna(g, "session_finished_timestamp"); col_lastmsg = achar_coluna(g, "last_message")
    col_lastts = achar_coluna(g, "last_message_timestamp")
    d = g[g[col_status] == "HumanHandover"].copy() if col_status else g.copy()
    d["duracao_h"] = (d[col_fim] - d[col_ini]).dt.total_seconds() / 3600
    d["horas_ociosas"] = (d[col_fim] - d[col_lastts]).dt.total_seconds() / 3600
    d["sem_resposta"] = d[col_lastmsg].astype(str).str.startswith(MSG_HANDOVER) if col_lastmsg else np.nan
    th = len(d); sr = int(d["sem_resposta"].sum()) if col_lastmsg else 0
    resumo = pd.DataFrame([
        ("Repassadas a vendedor", th), ("Encerradas sem resposta (inferido)", sr),
        ("% sem resposta", round(100 * sr / th, 1) if th else 0),
        ("Ociosidade mediana até fechar (h)", round(d["horas_ociosas"].median(), 1)),
    ], columns=["Métrica", "Valor"])
    return {"resumo_72h": resumo, "base_detalhada": d}


# =====================================================================
# GRÁFICOS
# =====================================================================

def _ax(ax, titulo):
    ax.set_title(titulo, fontsize=12.5, fontweight="bold", color=COR_NAVY, pad=10)
    for s in ["top", "right"]:
        ax.spines[s].set_visible(False)
    ax.tick_params(colors="#444", labelsize=8.5)
    ax.set_axisbelow(True)


def _save(fig, nome):
    os.makedirs(PASTA_GRAFICOS, exist_ok=True)
    p = os.path.join(PASTA_GRAFICOS, nome)
    fig.tight_layout(); fig.savefig(p, dpi=115, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return p


def _barh_rotulado(serie_idx, serie_val, titulo, nome, cor=COR_AZUL, sufixo="", figsize=(8, 5)):
    fig, ax = plt.subplots(figsize=figsize)
    barras = ax.barh([str(x) for x in serie_idx], serie_val, color=cor)
    ax.grid(axis="x", color="#E6EBF0", linewidth=0.8); _ax(ax, titulo)
    try:
        rotulos = [f"{(v if v == v else 0):g}{sufixo}" for v in serie_val]
        ax.bar_label(barras, labels=rotulos, padding=3, fontsize=8.5, color="#333")
    except Exception:
        pass
    return _save(fig, nome)


def gerar_graficos(g, v, at, ib, blocos):
    imgs = {}
    # ---------- ATENDIMENTO (núcleo) ----------
    if len(at):
        try:  # respondido vs não
            r = blocos["A1_resumo"]; dd = dict(zip(r["Métrica"], r["Valor"]))
            com = dd.get("Conversas respondidas por vendedor", 0); sem = dd.get("Conversas SEM resposta do vendedor", 0)
            fig, ax = plt.subplots(figsize=(6, 4.2))
            b = ax.bar(["Respondidas", "Sem resposta"], [com, sem], color=[COR_OK, COR_ALERTA])
            ax.grid(axis="y", color="#E6EBF0"); _ax(ax, "Conversas: respondidas x sem resposta do vendedor")
            ax.bar_label(b, padding=3, fontsize=10)
            imgs["A1_resumo"] = _save(fig, "A1_resposta.png")
        except Exception as e: print("[g A1]", e)

        try:  # 24h semanal — barras absolutas sobrepostas (atendidas ⊆ chegaram, mesma barra)
            t = blocos["B_24h_semanal"]
            fig, ax = plt.subplots(figsize=(9.5, 4.6))
            x = np.arange(len(t))
            b1 = ax.bar(x, t["chegaram"], 0.62, color=COR_CIANO, label="chegaram ao vendedor")
            b2 = ax.bar(x, t["respondidos_24h"], 0.34, color=COR_OK, label="conversas atendidas")
            _ax(ax, "Chegada ao vendedor x conversas atendidas (semanal)"); ax.set_ylabel("conversas")
            labels_sem = [f"Sem. {row['dia'].strftime('%d/%m')}" for _, row in t.iterrows()]
            ax.set_xticks(x); ax.set_xticklabels(labels_sem, rotation=45, ha="right")
            ax.bar_label(b1, padding=2, fontsize=8); ax.bar_label(b2, padding=2, fontsize=8)
            ax.legend(loc="upper right", fontsize=8)
            imgs["B_24h_semanal"] = _save(fig, "B_24h.png")
        except Exception as e: print("[g B]", e)

        try:  # 24h semanal — empilhado 100% normalizado: conversas atendidas + não atendidas
            t = blocos["B_24h_semanal"]
            chegaram_arr = t["chegaram"].values
            atendidos_arr = t["respondidos_24h"].values.astype(int)
            nao_at_arr = chegaram_arr - atendidos_arr
            labels_sem = [f"Sem. {row['dia'].strftime('%d/%m')}" for _, row in t.iterrows()]
            # Normaliza pela chegada total (atendidas ⊆ chegaram)
            pct_at  = np.where(chegaram_arr > 0, 100.0 * atendidos_arr / chegaram_arr, 0.0)
            pct_nao = np.where(chegaram_arr > 0, 100.0 * nao_at_arr   / chegaram_arr, 0.0)
            fig, ax = plt.subplots(figsize=(9.5, 4.6))
            x = np.arange(len(t)); bar_w = 0.6
            ax.bar(x, pct_at,  bar_w, color=COR_OK,    label="Conversas atendidas")
            ax.bar(x, pct_nao, bar_w, bottom=pct_at, color=COR_ALERTA, label="Conversas não atendidas")
            for i, (pa, pn) in enumerate(zip(pct_at, pct_nao)):
                if pa > 5:
                    ax.text(x[i], pa / 2, f"{pa:.0f}%", ha="center", va="center",
                            color="white", fontweight="bold", fontsize=10)
                if pn > 5:
                    ax.text(x[i], pa + pn / 2, f"{pn:.0f}%", ha="center", va="center",
                            color="white", fontweight="bold", fontsize=10)
            ax.set_xticks(x); ax.set_xticklabels(labels_sem, rotation=45, ha="right")
            ax.set_ylim(0, 108)
            ax.set_ylabel("% do total de chegadas"); ax.grid(axis="y", color="#E6EBF0")
            ax.legend(loc="upper right", fontsize=8)
            _ax(ax, "Conversas atendidas — distribuição proporcional por semana (%)")
            imgs["B_24h_stacked"] = _save(fig, "B_24h_stacked.png")
        except Exception as e: print("[g B_stacked]", e)

        try:  # last sender geral
            t = blocos["A2_last_sender"]
            fig, ax = plt.subplots(figsize=(6, 4.2))
            cores = {"IA": COR_CINZA, "AGENT": COR_AZUL, "USER": COR_AMBAR}
            b = ax.bar(t["last_message_sender"].astype(str),
                       t["conversas"], color=[cores.get(x, COR_AZUL) for x in t["last_message_sender"]])
            ax.grid(axis="y", color="#E6EBF0"); _ax(ax, "Quem enviou a última mensagem (geral)")
            ax.bar_label(b, padding=3, fontsize=10)
            imgs["A2_last_sender"] = _save(fig, "A2_sender.png")
        except Exception as e: print("[g A2]", e)

        try:  # finish type geral
            t = blocos["A3_finish_type"]
            fig, ax = plt.subplots(figsize=(5.5, 4.2))
            ax.pie(t["conversas"], labels=t["finish_type"].astype(str), colors=[COR_ALERTA, COR_OK, COR_CINZA][:len(t)],
                   autopct=lambda p: f"{p:.1f}%", startangle=90, wedgeprops=dict(edgecolor="white"))
            ax.set_title("Tipo de encerramento (AUTO x MANUAL)", fontsize=12.5, fontweight="bold", color=COR_NAVY)
            imgs["A3_finish_type"] = _save(fig, "A3_finish.png")
        except Exception as e: print("[g A3]", e)

    # ---------- POR VENDEDOR ----------
    pv = blocos.get("D1_por_vendedor")
    if pv is not None and len(pv):
        # desligados sem conversa recebida na janela ficam fora dos GRÁFICOS
        pv = pv[~(pv["vendedor"].isin(AGENTES_DESLIGADOS) & (pv["chats_recebidos"] == 0))]
    if pv is not None and len(pv):
        s = pv.sort_values("chats_recebidos")
        imgs["D1_por_vendedor"] = _barh_rotulado(s["vendedor"], s["chats_recebidos"],
                                                 "Conversas recebidas por vendedor", "D1_recebidos.png")
        # gráficos de PERCENTUAL só com vendedores que receberam conversa no
        # período: com 0 recebidas o % é trivial e distorce o ranking
        pv_pct = pv[pv["chats_recebidos"] > 0]
        s2 = pv_pct.sort_values("%_nao_respondidos")
        imgs["D1b_nao_resp"] = _barh_rotulado(s2["vendedor"], s2["%_nao_respondidos"],
                                              "% de conversas não respondidas por vendedor", "D2_naoresp.png",
                                              cor=COR_ALERTA, sufixo="%")
        # gráfico de % RESPONDIDOS (complemento — usado no layout do relatório)
        pv_resp = pv_pct.copy(); pv_resp["%_respondidos"] = round(100 - pv_resp["%_nao_respondidos"], 1)
        s2r = pv_resp.sort_values("%_respondidos")
        imgs["D1b_resp"] = _barh_rotulado(s2r["vendedor"], s2r["%_respondidos"],
                                          "% de conversas respondidas por vendedor", "D2_resp.png",
                                          cor=COR_OK, sufixo="%")
        s3 = pv_pct.sort_values("%_em_24h")
        imgs["D1c_24h"] = _barh_rotulado(s3["vendedor"], s3["%_em_24h"],
                                         "% de conversas atendidas por vendedor", "D3_24h.png", cor=COR_OK, sufixo="%")
        s4 = pv.dropna(subset=["tempo_med_1a_resp_h"]).sort_values("tempo_med_1a_resp_h", ascending=False)
        if len(s4):
            imgs["D1d_tempo"] = _barh_rotulado(s4["vendedor"], s4["tempo_med_1a_resp_h"],
                                               "Tempo mediano de 1ª resposta (h)", "D4_tempo.png", cor=COR_AMBAR)
        try:  # evolução semanal do tempo mediano de 1ª resposta do time (linha única)
            # duas variantes do mesmo gráfico: eixo y em horas e em minutos
            ts = tempo_resposta_semanal_time(at)
            if len(ts):
                blocos["D3_tempo_semanal"] = ts
                xs = list(range(len(ts)))
                labels_x = [pd.Timestamp(d).strftime("%d/%m") for d in ts["dia"]]
                for unidade, fator, sufixo in (("h", 1, ""), ("min", 60, "_min")):
                    vals = (ts["tempo_med_h"] * fator).round(2 if fator == 1 else 0)
                    fig, ax = plt.subplots(figsize=(10, 5.2))
                    ax.plot(xs, vals, marker="o", linewidth=2.2, markersize=6,
                            color=COR_AMBAR, zorder=3)
                    for x_, v_ in zip(xs, vals):
                        rotulo = f"{v_:.0f}" if fator == 60 else f"{v_:.2f}".replace(".", ",")
                        ax.annotate(rotulo, (x_, v_), textcoords="offset points",
                                    xytext=(0, 9), ha="center", fontsize=8.5,
                                    color=COR_NAVY, fontweight="bold")
                    ax.set_xticks(xs)
                    ax.set_xticklabels(labels_x, rotation=45, ha="right")
                    ax.set_ylim(0, float(vals.max()) * 1.2 if float(vals.max()) > 0 else 1)
                    ax.set_ylabel(f"Tempo mediano 1ª resposta ({unidade})")
                    ax.grid(axis="y", color="#E6EBF0", linewidth=0.8)
                    _ax(ax, f"Tempo mediano de 1ª resposta ({unidade}) — evolução semanal do time")
                    imgs["D3_tempo_semanal" + sufixo] = _save(fig, f"D3_tempo_semanal{sufixo}.png")
        except Exception as e: print("[g D3_semanal]", e)
        try:  # finish stacked
            fig, ax = plt.subplots(figsize=(8, 5)); s5 = pv.iloc[::-1]
            ax.barh(s5["vendedor"], s5["fim_AUTO"], color=COR_ALERTA, label="AUTO")
            ax.barh(s5["vendedor"], s5["fim_MANUAL"], left=s5["fim_AUTO"], color=COR_OK, label="MANUAL")
            ax.grid(axis="x", color="#E6EBF0"); _ax(ax, "Encerramento por vendedor (AUTO x MANUAL)")
            ax.legend(fontsize=8)
            imgs["D1e_finish"] = _save(fig, "D5_finish_vend.png")
        except Exception as e: print("[g D5]", e)
        try:  # sender stacked
            fig, ax = plt.subplots(figsize=(8, 5)); s6 = pv.iloc[::-1]
            ax.barh(s6["vendedor"], s6["ult_msg_IA"], color=COR_CINZA, label="IA")
            ax.barh(s6["vendedor"], s6["ult_msg_AGENT"], left=s6["ult_msg_IA"], color=COR_AZUL, label="Agent")
            ax.barh(s6["vendedor"], s6["ult_msg_USER"], left=s6["ult_msg_IA"] + s6["ult_msg_AGENT"],
                    color=COR_AMBAR, label="User")
            ax.grid(axis="x", color="#E6EBF0"); _ax(ax, "Última mensagem por vendedor (IA/Agent/User)")
            ax.legend(fontsize=8)
            imgs["D1f_sender"] = _save(fig, "D6_sender_vend.png")
        except Exception as e: print("[g D6]", e)

    tr = blocos.get("D2_transferencias")
    if tr is not None and len(tr):
        try:
            fig, ax = plt.subplots(figsize=(8, 5)); y = np.arange(len(tr)); h = 0.4
            ax.barh(y - h/2, tr["recebidos_via_transfer"], height=h, color=COR_AZUL, label="recebidas via transfer")
            ax.barh(y + h/2, tr["transferiu_out"], height=h, color=COR_AMBAR, label="transferiu (out)")
            ax.set_yticks(y); ax.set_yticklabels(tr["vendedor"]); ax.grid(axis="x", color="#E6EBF0")
            _ax(ax, "Transferências por vendedor (recebidas x enviadas)"); ax.legend(fontsize=8)
            imgs["D2_transferencias"] = _save(fig, "D7_transfer.png")
        except Exception as e: print("[g D7]", e)


    # ---------- PADRÃO DE RESÍDUO DOS CONTATOS ATRIBUÍDOS ----------
    rc_comp = blocos.get("R1_residuo_comp")
    if rc_comp is not None and len(rc_comp):
        try:  # pizza de composição por padrão de resíduo
            fig, ax = plt.subplots(figsize=(6.6, 4.8))
            paleta = [COR_ALERTA, COR_OK, COR_AMBAR, COR_AZUL, COR_CIANO, COR_CINZA, COR_NAVY]
            cores = [paleta[i % len(paleta)] for i in range(len(rc_comp))]
            ax.pie(rc_comp["contatos"], labels=rc_comp["padrao"].astype(str), colors=cores,
                   autopct=lambda p: f"{p:.0f}%", startangle=90, wedgeprops=dict(edgecolor="white"),
                   textprops={"fontsize": 8})
            ax.set_title("Composição dos contatos por padrão de resíduo",
                         fontsize=12.5, fontweight="bold", color=COR_NAVY)
            imgs["R1_residuo_comp"] = _save(fig, "R1_residuo_comp.png")
        except Exception as e: print("[g R1]", e)

    rc_resp = blocos.get("R2_residuo_resp")
    if rc_resp is not None and len(rc_resp):
        try:  # barras horizontais empilhadas com % dentro de cada segmento
            t = rc_resp.sort_values("contatos").reset_index(drop=True)
            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            ax.barh(t["padrao"].astype(str), t["respondidos"], color=COR_OK, label="respondidos")
            ax.barh(t["padrao"].astype(str), t["sem_resposta"], left=t["respondidos"],
                    color=COR_ALERTA, label="sem resposta")
            for i, row in t.iterrows():
                total = row["contatos"]
                if total > 0:
                    pct_r = row["respondidos"] / total * 100
                    pct_s = row["sem_resposta"] / total * 100
                    if row["respondidos"] > 0 and pct_r >= 8:
                        ax.text(row["respondidos"] / 2, i, f"{pct_r:.0f}%",
                                va="center", ha="center", color="white", fontsize=8)
                    if row["sem_resposta"] > 0 and pct_s >= 8:
                        ax.text(row["respondidos"] + row["sem_resposta"] / 2, i,
                                f"{pct_s:.0f}%",
                                va="center", ha="center", color="white", fontsize=8)
            ax.grid(axis="x", color="#E6EBF0")
            _ax(ax, "Respondidos x sem resposta por padrão de resíduo (volume)")
            ax.legend(loc="lower right", fontsize=8)
            imgs["R2_residuo_resp"] = _save(fig, "R2_residuo_resp.png")
        except Exception as e: print("[g R2]", e)

    # ---------- RETORNO ----------
    rc = blocos.get("C1_retorno_dist")
    if rc is not None and len(rc):
        try:
            fig, ax = plt.subplots(figsize=(6.5, 4.2))
            b = ax.bar(rc["faixa"].astype(str), rc["contatos"], color=COR_AZUL)
            ax.grid(axis="y", color="#E6EBF0"); _ax(ax, "Distribuição de conversas por contato (retorno)")
            ax.bar_label(b, padding=3, fontsize=9)
            imgs["C1_retorno_dist"] = _save(fig, "C1_retorno.png")
        except Exception as e: print("[g C1]", e)

    # ---------- INBOX ----------
    if len(ib):
        try:
            t = blocos["E1_inbox"]
            fig, ax = plt.subplots(figsize=(9.5, 4.4))
            if "total_recebidos" in t:
                ax.bar(t["data"], t["total_recebidos"], color=COR_CIANO, label="bate-papos recebidos")
            if "contatos_distintos" in t:
                ax.plot(t["data"], t["contatos_distintos"], color=COR_NAVY, marker="o", markersize=3,
                        linewidth=1.5, label="contatos distintos")
            _ax(ax, "Inbox: volume diário recebido x contatos distintos"); ax.grid(axis="y", color="#E6EBF0")
            ax.legend(fontsize=8); fig.autofmt_xdate(rotation=45)
            imgs["E1_inbox"] = _save(fig, "E1_inbox.png")
        except Exception as e: print("[g E1]", e)

    # ---------- ANTERIORES ----------
    if len(g):
        try:
            st = blocos.get("_status_geral")
            if st and sum(st.values()):
                fig, ax = plt.subplots(figsize=(6, 4.2)); vals = [st.get("AI", 0), st.get("HumanHandover", 0)]
                ax.pie(vals, labels=["Parou no bot (AI)", "Repassada"], colors=[COR_CIANO, COR_AZUL],
                       autopct=lambda p: f"{p:.1f}%", startangle=90, wedgeprops=dict(width=0.42, edgecolor="white"))
                ax.set_title("Destino das sessões (log Treble)", fontsize=12.5, fontweight="bold", color=COR_NAVY)
                imgs["F1_status"] = _save(fig, "F1_status.png")
        except Exception as e: print("[g F1]", e)
        try:
            f = blocos.get("F2_funil")
            if f is not None and len(f):
                top = f.head(15).iloc[::-1]; ai = top.get("AI", 0); hh = top.get("HumanHandover", 0)
                fig, ax = plt.subplots(figsize=(8, 5)); yy = range(len(top))
                ax.barh(yy, ai, color=COR_CIANO, label="AI"); ax.barh(yy, hh, left=ai, color=COR_AZUL, label="Handover")
                ax.set_yticks(list(yy)); ax.set_yticklabels(top[top.columns[0]].astype(str), fontsize=8)
                _ax(ax, "Onde as conversas terminam (top 15 nós)"); ax.grid(axis="x", color="#E6EBF0"); ax.legend(fontsize=8)
                imgs["F2_funil"] = _save(fig, "F2_funil.png")
        except Exception as e: print("[g F2]", e)
        try:
            t = blocos.get("F3_tendencia")
            if t is not None and len(t):
                fig, ax = plt.subplots(figsize=(9, 4.4)); ax.bar(t["periodo"], t["sessoes"], color=COR_CIANO)
                _ax(ax, "Volume diário e taxa de repasse"); ax.set_ylabel("sessões")
                if "taxa_repasse_%" in t:
                    a2 = ax.twinx(); a2.plot(t["periodo"], t["taxa_repasse_%"], color=COR_ALERTA, marker="o", markersize=3)
                    a2.set_ylabel("% repasse", color=COR_ALERTA); a2.tick_params(axis="y", colors=COR_ALERTA); a2.spines["top"].set_visible(False)
                fig.autofmt_xdate(rotation=45)
                imgs["F3_tendencia"] = _save(fig, "F3_tend.png")
        except Exception as e: print("[g F3]", e)

    if len(v):
        try:
            c = blocos.get("G1_caminhos")
            if c is not None and len(c):
                fig, ax = plt.subplots(figsize=(6.5, 4.4))
                ax.pie(c["conversas"], labels=c["caminho"], colors=[COR_AZUL, COR_CIANO, COR_OK, COR_CINZA][:len(c)],
                       autopct=lambda p: f"{p:.1f}%", startangle=90, textprops=dict(fontsize=8), wedgeprops=dict(edgecolor="white"))
                ax.set_title("Distribuição por caminho do fluxo", fontsize=12.5, fontweight="bold", color=COR_NAVY)
                imgs["G1_caminhos"] = _save(fig, "G1_caminhos.png")
        except Exception as e: print("[g G1]", e)
        try:
            comp = blocos.get("G2_completude")
            if comp is not None and len(comp):
                cc = comp.sort_values("%_preenchido")
                imgs["G2_completude"] = _barh_rotulado(cc["campo"], cc["%_preenchido"],
                                                       "Completude dos campos de qualificação (%)", "G2_comp.png", sufixo="%")
        except Exception as e: print("[g G2]", e)
        # completude por caminho: "Sou novo por aqui"
        try:
            comp_n = blocos.get("G2b_completude_novo")
            n_n = blocos.get("_n_sessoes_novo", 0)
            if comp_n is not None and len(comp_n):
                # ordem REAL do fluxo (aba já vem ordenada); invertida para a
                # 1ª pergunta ficar no TOPO do gráfico horizontal
                cc = comp_n.iloc[::-1]
                imgs["G2b_completude_novo"] = _barh_rotulado(cc["campo"], cc["%_preenchido"],
                    f"Preenchimento dos campos — Sou novo por aqui ({n_n} sessões)",
                    "G2b_novo.png", sufixo="%")
        except Exception as e: print("[g G2b]", e)
        # completude por caminho: "Já sou cliente"
        try:
            comp_c = blocos.get("G2c_completude_cli")
            n_c = blocos.get("_n_sessoes_cli", 0)
            if comp_c is not None and len(comp_c):
                # ordem REAL do fluxo, 1ª pergunta no topo (ver G2b acima)
                cc = comp_c.iloc[::-1]
                imgs["G2c_completude_cli"] = _barh_rotulado(cc["campo"], cc["%_preenchido"],
                    f"Preenchimento dos campos — Já sou cliente ({n_c} sessões)",
                    "G2c_cli.png", sufixo="%")
        except Exception as e: print("[g G2c]", e)
        try:
            pe = blocos.get("G3_por_estado")
            if pe is not None and len(pe):
                pp = pe.dropna(subset=["UF"]).head(12).sort_values("conversas")
                imgs["G3_por_estado"] = _barh_rotulado(pp["UF"], pp["conversas"],
                                                       "Conversas por estado (normalizado)", "G3_estado.png")
        except Exception as e: print("[g G3]", e)

    if len(g):
        try:
            r = blocos.get("F4_72h")
            if r is not None and len(r):
                dd = dict(zip(r["Métrica"], r["Valor"]))
                total = dd.get("Repassadas a vendedor", 0); sem = dd.get("Encerradas sem resposta (inferido)", 0)
                fig, ax = plt.subplots(figsize=(6, 4.2))
                b = ax.bar(["Com resposta", "Sem resposta"], [max(total - sem, 0), sem], color=[COR_OK, COR_ALERTA])
                ax.grid(axis="y", color="#E6EBF0"); _ax(ax, "SLA 72h (log de sessões): resposta x sem resposta")
                ax.bar_label(b, padding=3, fontsize=10)
                imgs["F4_72h"] = _save(fig, "F4_72h.png")
        except Exception as e: print("[g F4]", e)
    return imgs


# =====================================================================
# EXPORTAÇÃO — PAINEL (1ª aba) + tabelas com gráfico embutido
# =====================================================================

# ordem narrativa dos gráficos no painel: (rótulo, chave_imagem)
ORDEM_PAINEL = [
    # 01 — CONVERSAS
    ("Conversas: respondidas x sem resposta", "A1_resumo"),
    # 02 — CONVERSAS ATENDIDAS
    ("Chegada ao vendedor x conversas atendidas (semanal)", "B_24h_semanal"),
    ("Conversas atendidas — distribuição proporcional (%)", "B_24h_stacked"),
    # 03 — VOLUME POR VENDEDOR
    ("Conversas recebidas por vendedor", "D1_por_vendedor"),
    ("% de conversas respondidas por vendedor", "D1b_resp"),
    # 04 — TEMPO DE RESPOSTA
    ("Tempo mediano de 1ª resposta (h) — evolução semanal do time", "D3_tempo_semanal"),
    ("Tempo mediano de 1ª resposta (h) — resumo do período", "D1d_tempo"),
    # 05 — PADRÃO DE RESÍDUO
    ("Composição dos contatos por padrão de resíduo", "R1_residuo_comp"),
    ("Respondidos x sem resposta por padrão de resíduo", "R2_residuo_resp"),
    # 06 — QUALIFICAÇÃO
    ("Preenchimento dos campos — Sou novo por aqui", "G2b_completude_novo"),
    ("Preenchimento dos campos — Já sou cliente", "G2c_completude_cli"),
]


def exportar_resultados(blocos, imagens, caminho=ARQUIVO_SAIDA):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment

    with pd.ExcelWriter(caminho, engine="openpyxl") as w:
        # 1) escreve as abas de tabela (na ordem do dict)
        for nome, df in blocos.items():
            if nome.startswith("_") or not isinstance(df, pd.DataFrame) or not len(df):
                continue
            aba = str(nome)[:31]
            df.to_excel(w, sheet_name=aba, index=False)
            png = imagens.get(nome)
            if png and os.path.exists(png):
                w.sheets[aba].add_image(XLImage(png), f"A{len(df) + 4}")

        wb = w.book
        # 2) cria o PAINEL como primeira aba
        ws = wb.create_sheet("00_PAINEL", 0)
        ws.sheet_view.showGridLines = False
        ws["A1"] = "RELATÓRIO — TREBLE"
        ws["A1"].font = Font(bold=True, size=16, color="0F2A47")
        _ini = globals().get("_SEMANA_INI"); _fim = globals().get("_SEMANA_FIM")
        _per = f"  ·  Semana de {_ini:%d/%m/%Y} a {_fim:%d/%m/%Y}" if _ini is not None else ""
        ws["A2"] = "Nexforce · RevOps — leitura geral da operação (base: atendimento-treble + log de sessões)" + _per
        ws["A2"].font = Font(size=10, color="5A6472")

        ws.column_dimensions["A"].width = 14
        LARG = 13  # nº de colunas que as faixas/títulos abrangem
        linha = 4
        for rotulo, chave in ORDEM_PAINEL:
            if chave is None:  # faixa de seção (mesclada e destacada)
                ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=LARG)
                cell = ws.cell(row=linha, column=1, value=rotulo.strip(" —"))
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F5C99")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[linha].height = 22
                linha += 2
                continue
            png = imagens.get(chave)
            if not (png and os.path.exists(png)):
                continue
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=LARG)
            t = ws.cell(row=linha, column=1, value=rotulo)
            t.font = Font(bold=True, size=10, color="0F2A47")
            ws.add_image(XLImage(png), f"A{linha + 1}")
            linha += 24

    print(f"\n[ok] Excel: {caminho}  (1ª aba = 00_PAINEL)")
    print(f"[ok] PNGs em ./{PASTA_GRAFICOS}/")


def _letra(col):
    from openpyxl.utils import get_column_letter
    return get_column_letter(col)


# =====================================================================
# RUNNER
# =====================================================================


# =====================================================================
# FILTRO SEMANAL — domingo a sábado da semana anterior
# =====================================================================
def janela_semana(at, g):
    """Resolve a janela (inicio, fim) a ser analisada, nesta ordem de prioridade:
       1) SEMANA_INICIO/SEMANA_FIM explícitos;
       2) detecção automática pelo intervalo das datas presentes nos dados;
       3) semana anterior (domingo a sábado) relativa a DATA_REFERENCIA/hoje."""
    if SEMANA_INICIO and SEMANA_FIM:
        ini = pd.Timestamp(SEMANA_INICIO)
        fim = pd.Timestamp(SEMANA_FIM) + pd.Timedelta(hours=23, minutes=59, seconds=59)
        return ini, fim
    # tenta detectar pelos dados (sessões e atendimento)
    datas = []
    for df, cols in [(g, ["session_started_timestamp", "first_message_timestamp"]),
                     (at, ["created_at", "assigned_at"])]:
        if df is not None and len(df):
            for c in cols:
                real = achar_coluna(df, c)
                if real is not None:
                    datas.append(pd.to_datetime(df[real], errors="coerce"))
                    break
    if datas:
        serie = pd.concat(datas)
        if serie.notna().any():
            return serie.min().normalize(), serie.max().normalize() + pd.Timedelta(hours=23, minutes=59, seconds=59)
    return janela_semana_anterior(DATA_REFERENCIA)


def janela_semana_anterior(ref=None):
    """Retorna (domingo, sabado_23h59) da semana imediatamente anterior à
    semana que contém `ref`. Semana definida de DOMINGO a SÁBADO."""
    ref = pd.Timestamp(ref).normalize() if ref else pd.Timestamp.today().normalize()
    dow = (ref.weekday() + 1) % 7            # domingo=0 ... sábado=6
    domingo_semana_atual = ref - pd.Timedelta(days=int(dow))
    sabado = domingo_semana_atual - pd.Timedelta(days=1)
    domingo = sabado - pd.Timedelta(days=6)
    return domingo, sabado + pd.Timedelta(hours=23, minutes=59, seconds=59)


def _filtra_periodo(df, cols, ini, fim):
    """Filtra df pela 1ª coluna de data existente dentre `cols`, no intervalo."""
    if df is None or not len(df):
        return df
    for c in cols:
        real = achar_coluna(df, c)
        if real is not None:
            s = pd.to_datetime(df[real], errors="coerce")
            return df[(s >= ini) & (s <= fim)].copy()
    return df


def main(caminho=ARQUIVO):
    dados = carregar_arquivo(caminho)
    g = consolidar_gerais(dados)
    v = consolidar_versoes(dados)
    at = consolidar_atendimento(dados)
    ib = consolidar_inbox(dados)

    # ---- FILTRO DA SEMANA (domingo a sábado) ----
    ini, fim = janela_semana(at, g)
    globals()["_SEMANA_INI"], globals()["_SEMANA_FIM"] = ini, fim
    print(f"\n>>> Janela analisada: {ini:%d/%m/%Y} (dom) a {fim:%d/%m/%Y} (sáb)")
    at = _filtra_periodo(at, ["created_at", "assigned_at"], ini, fim)
    g  = _filtra_periodo(g, ["session_started_timestamp", "first_message_timestamp", "session_finished_timestamp"], ini, fim)
    if len(ib):
        ib = _filtra_periodo(ib, ["data"], ini, fim)
    # versões (qualificação) são filtradas pela PRÓPRIA data do registro,
    # pois refletem TODOS os contatos do fluxo na semana (não só os que
    # chegaram a um vendedor).
    if len(v):
        v = _filtra_periodo(v, ["Data", "ultima atividade", "ultima_atividade_dt"], ini, fim)
    print(f">>> Após filtro — atendimento: {len(at)} | sessões: {len(g)} | versões: {len(v)} | inbox: {len(ib)}\n")

    blocos = {}

    # ---- A/B/C/D: ATENDIMENTO (base de verdade) ----
    if len(at):
        blocos["A1_resumo"] = atendimento_geral(at)
        blocos["A0_panorama"] = panorama_time(at)
        blocos["B_24h_diario"] = resposta_24h_diaria(at)
        blocos["B_24h_semanal"] = resposta_24h_semanal(at)
        ret = retorno_clientes(at)
        blocos["C0_retorno_resumo"] = ret["resumo_retorno"]
        blocos["C1_retorno_dist"] = ret["distribuicao"]
        blocos["C2_recorrentes"] = ret["recorrentes"]
        blocos["D1_por_vendedor"] = por_vendedor(at)
        blocos["D2_transferencias"] = transferencias(at)
        fichas = {}
    else:
        fichas = {}

    # ---- Padrão de resíduo dos contatos atribuídos (atendimento x versões) ----
    if len(at) and len(v):
        pr = padrao_residuo_atribuidos(at, v)
        if pr:
            blocos["R1_residuo_comp"] = pr["composicao"]
            blocos["R2_residuo_resp"] = pr["resposta_por_padrao"]

    # ---- E: INBOX ----
    if len(ib):
        blocos["E1_inbox"] = inbox_diario(ib)

    # ---- F: log de sessões (mantido) ----
    if len(g):
        vg = visao_geral_periodo(g)
        blocos["F0_resumo_sessoes"] = vg["resumo"]; blocos["_status_geral"] = vg["_status"]
        blocos["F2_funil"] = funil_por_no(g)
        blocos["F3_tendencia"] = tendencia_temporal(g, "D")

    # ---- G: abas de versão (mantido) ----
    if len(v):
        cam = analise_caminhos(v)
        blocos["G1_caminhos"] = cam["distribuicao_caminho"]
        blocos["G2_completude"] = cam["completude_qualificacao"]
        blocos["G2b_completude_novo"] = cam["completude_novo"]
        blocos["G2c_completude_cli"] = cam["completude_cli"]
        blocos["_n_sessoes_novo"] = cam["n_sessoes_novo"]
        blocos["_n_sessoes_cli"] = cam["n_sessoes_cli"]
        blocos["_n_sessoes_total"] = cam["n_sessoes_total"]
        q = qualidade_dados(v)
        blocos["G4_qualidade"] = q["flags_qualidade"]
        blocos["G5_estados_orig"] = q["estado_valores_originais"]
        terr = analise_territorial(v)
        blocos["G3_por_estado"] = terr.get("por_estado", pd.DataFrame())
        blocos["G6_por_cidade"] = terr.get("por_cidade", pd.DataFrame())

    imagens = gerar_graficos(g, v, at, ib, blocos)

    # gráfico individual de cada vendedor (embutido na própria aba)
    for nome_vend, df_ficha in fichas.items():
        p = grafico_ficha(nome_vend, df_ficha)
        if p:
            imagens[nome_vend] = p

    # prints de conferência
    for chave in ["A1_resumo", "D1_por_vendedor", "C0_retorno_resumo"]:
        if chave in blocos:
            print(f"\n===== {chave} =====")
            print(blocos[chave].to_string(index=False))

    exportar_resultados(blocos, imagens)
    return blocos


if __name__ == "__main__":
    import argparse as _ap
    _parser = _ap.ArgumentParser(
        description='Análise de equipe por período. Exemplo: python scripts/geral/analise_equipe_periodo.py "01-01-2026 a 30-06-2026"'
    )
    _parser.add_argument(
        'periodo', nargs='?', default=None,
        metavar='DD-MM-AAAA a DD-MM-AAAA',
        help='Período da análise. Define ARQUIVO, SEMANA_INICIO e SEMANA_FIM automaticamente.'
    )
    _args = _parser.parse_args()

    if _args.periodo:
        import re as _re
        _m = _re.match(
            r'(\d{2})-(\d{2})-(\d{4})\s+a\s+(\d{2})-(\d{2})-(\d{4})',
            _args.periodo.strip()
        )
        if not _m:
            print(f'Formato inválido: "{_args.periodo}". Use DD-MM-AAAA a DD-MM-AAAA')
            import sys; sys.exit(1)
        d1, m1, y1, d2, m2, y2 = _m.groups()
        SEMANA_INICIO = f'{y1}-{m1}-{d1}'
        SEMANA_FIM    = f'{y2}-{m2}-{d2}'
        # pasta canônica do período primeiro; fallback para o caminho literal
        _cand = f'analises/{_args.periodo.strip()} (geral)/Dados_Treble_Semana.xlsx'
        ARQUIVO = _cand if os.path.exists(_cand) else f'{_args.periodo}/Dados_Treble_Semana.xlsx'
        print(f'[CLI] Período: {_args.periodo} | Arquivo: {ARQUIVO}')

    main(ARQUIVO)