#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_relatorio_html.py
Gera o relatório de atendimento Treble em HTML (dark theme).

PNGs gerados pelo analise_treble_semanal.py são embutidos como base64 (arquivo autocontido).

Nexforce · RevOps
"""

import os
import base64
from datetime import datetime

import openpyxl

# ═══════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO  — sobrescrito via CLI (--tipo, --xlsx, --graficos,
#                 --d3-unidade, --com-titulos)
# ═══════════════════════════════════════════════════════════════════════
# Nome do cliente exibido no título/capa do relatório. Ajuste para o seu.
CLIENTE       = "Cliente"
# Padrão de formato: gráfico de evolução em HORAS e SEM os captions cinza.
TIPO          = 'semanal'
SEMANA        = None
XLSX_ANALISE  = "relatorio_treble_semanal.xlsx"
GRAFICOS      = "graficos/treble"
D3_UNIDADE    = 'h'      # 'h' (padrão aprovado) ou 'min' — eixo y da evolução semanal (S.04)
SEM_TITULOS   = True     # padrão aprovado: sem captions cinza acima dos gráficos


# ═══════════════════════════════════════════════════════════════════════
# CARGA DE DADOS
# ═══════════════════════════════════════════════════════════════════════

def load_data():
    wb = openpyxl.load_workbook(XLSX_ANALISE, read_only=True)

    def kv(name):
        if name not in wb.sheetnames:
            return {}
        ws = wb[name]
        d = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row[0] is not None:
                d[str(row[0])] = row[1]
        return d

    def table(name):
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        hdr = None
        out = []
        for row in ws.iter_rows(values_only=True):
            if hdr is None:
                hdr = [str(c) if c is not None else '' for c in row]
            else:
                out.append(dict(zip(hdr, row)))
        return out

    d = {
        'resumo':           kv('A1_resumo'),
        'sessoes':          kv('F0_resumo_sessoes'),
        'vendedores':       table('D1_por_vendedor'),
        'res_comp':         table('R1_residuo_comp'),
        'res_resp':         table('R2_residuo_resp'),
        'caminhos':         table('G1_caminhos'),
        'comp_novo':        table('G2b_completude_novo'),
        'comp_cli':         table('G2c_completude_cli'),
        'semanal_24h':      table('B_24h_semanal'),
        'diario_24h':       table('B_24h_diario'),
        'd3_tempo_semanal': table('D3_tempo_semanal'),
    }
    wb.close()
    return d


# ═══════════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ═══════════════════════════════════════════════════════════════════════

def pct(v, d=1):
    if v is None:
        return '—'
    return f'{float(v):.{d}f}%'.replace('.', ',')

def num_fmt(v, d=2):
    if v is None:
        return '—'
    return f'{float(v):.{d}f}'.replace('.', ',')

def img_b64(path):
    if not path or not os.path.exists(path):
        return ''
    with open(path, 'rb') as f:
        return base64.b64encode(f.read()).decode()

def g(name):
    return os.path.join(GRAFICOS, name)

_semana_cache = None
def semana_datas():
    global _semana_cache
    if _semana_cache:
        return _semana_cache
    s = SEMANA or '14-06-2026 a 20-06-2026'
    parts = s.replace(' a ', '_').split('_')
    def fmt(p):
        p = p.strip()
        pieces = p.replace('/', '-').split('-')
        if len(pieces) == 3:
            return f'{pieces[0]}/{pieces[1]}/{pieces[2]}'
        return p
    _semana_cache = (fmt(parts[0]), fmt(parts[1]))
    return _semana_cache


# ═══════════════════════════════════════════════════════════════════════
# CSS — dark theme
# ═══════════════════════════════════════════════════════════════════════

CSS = """
:root {
  --bg: #000000;
  --surface: #111111;
  --surface2: #1A1A1A;
  --border: #2A2A2A;
  --text: #FFFFFF;
  --muted: #888888;
  --prose: #C8C5BE;
  --accent: #1A6FFF;
  --accent2: #4D9FFF;
}

* { box-sizing: border-box; margin: 0; padding: 0; }

body {
  font-family: 'Lato', -apple-system, BlinkMacSystemFont, Arial, sans-serif;
  background: var(--bg);
  color: var(--text);
  font-size: 15px;
  line-height: 1.6;
}

.doc { max-width: 860px; margin: 0 auto; padding: 0 32px 80px; }

/* ── COVER ── */
.cover {
  min-height: 100vh;
  display: flex;
  flex-direction: column;
  justify-content: center;
  padding: 64px 0;
  border-bottom: 1px solid var(--border);
  margin-bottom: 64px;
}
.cover .eyebrow {
  font-family: 'JetBrains Mono', 'Courier New', monospace;
  font-size: 11px;
  letter-spacing: 0.12em;
  color: var(--accent);
  text-transform: uppercase;
  margin-bottom: 36px;
}
.cover h1 { font-size: 52px; font-weight: 700; line-height: 1.1; margin-bottom: 12px; }
.cover h1 span { color: var(--accent); }
.cover .subtitle { color: var(--muted); font-size: 14px; margin-bottom: 36px; }
.meta-row {
  display: flex; gap: 32px; flex-wrap: wrap;
  border-top: 1px solid var(--border);
  padding-top: 28px; margin-top: 28px;
}
.meta-item .label {
  font-family: 'JetBrains Mono', 'Courier New', monospace;
  font-size: 10px; letter-spacing: 0.1em;
  text-transform: uppercase; color: var(--muted);
  margin-bottom: 4px;
}
.meta-item .val { font-size: 14px; font-weight: 600; color: var(--text); }

/* ── STAT CARDS ── */
.stat-grid { display: flex; gap: 12px; margin: 24px 0; flex-wrap: wrap; }
.stat-card {
  flex: 1; min-width: 160px;
  background: var(--surface);
  border: 1px solid var(--border);
  border-radius: 8px;
  padding: 20px 18px 16px;
}
.stat-card .lbl {
  font-family: 'JetBrains Mono', 'Courier New', monospace;
  font-size: 10px; letter-spacing: 0.08em;
  text-transform: uppercase; color: var(--muted);
  margin-bottom: 8px;
  white-space: pre-line;
}
.stat-card .val { font-size: 32px; font-weight: 700; color: var(--accent); line-height: 1; }
.stat-card .sub { font-size: 11px; color: var(--muted); margin-top: 6px; }

/* ── SECTIONS ── */
.section { margin-top: 64px; }
.section-index {
  font-family: 'JetBrains Mono', 'Courier New', monospace;
  font-size: 11px; color: var(--accent);
  letter-spacing: 0.1em; text-transform: uppercase;
  margin-bottom: 10px;
}
.section-title { font-size: 28px; font-weight: 700; margin-bottom: 12px; }
.prose { font-size: 14px; color: var(--prose); line-height: 1.7; margin-bottom: 20px; }

/* ── SOURCE CAPTION ── */
.caption {
  background: var(--surface2);
  border-radius: 4px;
  padding: 7px 12px;
  font-family: 'JetBrains Mono', 'Courier New', monospace;
  font-size: 11px; color: var(--muted);
  margin-bottom: 12px;
}

/* ── CHART CARD ── */
.chart-img {
  background: #fff;
  border-radius: 6px;
  padding: 12px;
  margin-bottom: 24px;
  width: 100%;
}
.chart-img img { width: 100%; display: block; }

/* ── CALLOUT (leitura) ── */
.callout {
  background: var(--surface);
  border-left: 3px solid var(--accent);
  border-radius: 4px;
  padding: 16px 20px;
  margin: 20px 0;
  font-size: 14px;
  color: var(--text);
  line-height: 1.7;
}
.callout strong { color: var(--accent); }

/* ── CALLOUT-NOTE ── */
.callout-note {
  background: var(--surface2);
  border-left: 3px solid var(--muted);
  border-radius: 4px;
  padding: 12px 16px;
  margin: 12px 0 20px;
  font-size: 12px;
  color: var(--muted);
  font-family: 'JetBrains Mono', 'Courier New', monospace;
}

/* ── DATA TABLE ── */
.data-table { width: 100%; border-collapse: collapse; margin: 16px 0 24px; font-size: 13px; }
.data-table th {
  background: var(--surface2);
  color: var(--muted);
  font-family: 'JetBrains Mono', 'Courier New', monospace;
  font-size: 10px; letter-spacing: 0.08em; text-transform: uppercase;
  padding: 10px 12px; text-align: left;
  border-bottom: 1px solid var(--border);
}
.data-table td {
  padding: 10px 12px;
  border-bottom: 1px solid var(--border);
  color: var(--prose);
}
.data-table tr:nth-child(even) td { background: var(--surface); }
.data-table .num { color: var(--accent2); }

/* ── CLOSING ── */
.closing {
  margin-top: 80px;
  padding-top: 32px;
  border-top: 1px solid var(--border);
  font-size: 13px; color: var(--muted);
}
.closing .eyebrow {
  font-family: 'JetBrains Mono', 'Courier New', monospace;
  font-size: 11px; color: var(--accent);
  letter-spacing: 0.1em; text-transform: uppercase;
  margin-bottom: 16px;
}
"""


# ═══════════════════════════════════════════════════════════════════════
# BUILDERS DE COMPONENTES
# ═══════════════════════════════════════════════════════════════════════

def chart_tag(png_name, alt=''):
    path = g(png_name)
    b64 = img_b64(path)
    if not b64:
        return (f'<p style="color:var(--muted);font-size:12px;padding:12px">'
                f'[gráfico não encontrado: {png_name}]</p>')
    return (f'<div class="chart-img">'
            f'<img src="data:image/png;base64,{b64}" alt="{alt}">'
            f'</div>')

def section_header_html(num, name, desc):
    return (f'<div class="section-index">{num}   —   {name.upper()}</div>\n'
            f'<p class="prose">{desc}</p>\n')

def caption_html(text):
    # títulos cinza acima dos gráficos; fora do padrão aprovado (03/07/2026),
    # só entram com --com-titulos
    return '' if SEM_TITULOS else f'<div class="caption">{text}</div>'

def callout_html(text):
    return f'<div class="callout"><strong>Leitura:</strong> {text}</div>\n'

def callout_note_html(text):
    return f'<div class="callout-note">{text}</div>\n'

def stat_cards_html(cards):
    items = ''
    for lbl, val, sub in cards:
        items += (f'<div class="stat-card">'
                  f'<div class="lbl">{lbl}</div>'
                  f'<div class="val">{val}</div>'
                  f'<div class="sub">{sub}</div>'
                  f'</div>')
    return f'<div class="stat-grid">{items}</div>\n'

def summary_table_html(rows):
    trs = ''
    for key, val in rows:
        val_str = str(val) if val is not None else ''
        tok = val_str.split()[0] if val_str.split() else ''
        is_num = any(c in tok for c in ['%', 'h']) or tok.replace(',', '').replace('.', '').isdigit()
        cls = ' class="num"' if is_num else ''
        trs += f'<tr><td><strong>{key}</strong></td><td{cls}>{val_str}</td></tr>'
    return f'<table class="data-table"><tbody>{trs}</tbody></table>\n'

def residuo_table_html(data):
    # regra absoluta (06/07/2026): sem informação equivalente repetida —
    # "% sem resposta" é o complemento de "% respondida" e foi removida da tabela
    rows_html = ''
    for row in data:
        padrao = row.get('padrao', '')
        contatos = int(row.get('contatos') or 0)
        resp_pct = pct(row.get('%_respondida', 0))
        t_med = row.get('tempo_mediano_resp')
        t_str = f'{num_fmt(t_med, 2)} h' if t_med is not None and str(t_med) not in ('', 'None', 'nan') else '—'
        rows_html += (f'<tr><td>{padrao}</td>'
                      f'<td class="num">{contatos}</td>'
                      f'<td class="num">{resp_pct}</td>'
                      f'<td class="num">{t_str}</td></tr>')
    return (f'<table class="data-table">'
            f'<thead><tr>'
            f'<th>Padrão de resíduo</th>'
            f'<th>Contatos</th>'
            f'<th>% respondida</th>'
            f'<th>Tempo mediano de resposta</th>'
            f'</tr></thead>'
            f'<tbody>{rows_html}</tbody>'
            f'</table>\n')


# ═══════════════════════════════════════════════════════════════════════
# COVER (capa + sumário executivo)
# ═══════════════════════════════════════════════════════════════════════

def build_cover(d, semana_str):
    r = d['resumo']
    s = d['sessoes']
    conversas      = int(r.get('Conversas atendidas (total)', 0) or 0)
    sessoes_n      = int(s.get('Sessões no período', 0) or 0)
    pct_24h        = float(r.get('% respondidas em 24h (do total)', 0) or 0)
    tempo          = r.get('Tempo mediano de 1ª resposta (h)', 0) or 0
    resp_24h       = int(r.get('Respondidas dentro de 24h', 0) or 0)

    sem_ini, sem_fim = semana_datas()
    sem_display = f'{sem_ini} a {sem_fim}'

    # regra absoluta (06/07/2026): sem informação equivalente repetida —
    # "% de atendimento realizado" era equivalente a "% de conversas atendidas"
    # e foi removido dos cards e da tabela do sumário
    cards = stat_cards_html([
        ('% de conversas\natendidas',
         pct(pct_24h, 1),
         f'{resp_24h} de {conversas} conversas'),
        ('tempo mediano de\n1ª resposta',
         f'{num_fmt(tempo, 2)} h',
         'apenas conversas com resposta'),
    ])

    # top 3 pelo MAIOR % de resposta (desempate: mais respondidos)
    vendedores_resp = sorted(
        [v for v in d['vendedores'] if (v.get('respondidos') or 0) > 0],
        key=lambda v: (-(100 - (v.get('%_nao_respondidos') or 0)),
                       -(v.get('respondidos') or 0))
    )
    top_vendors = ', '.join(
        f"{v['vendedor']} ({pct(100 - (v.get('%_nao_respondidos') or 0), 0)})"
        for v in vendedores_resp[:3]
    )
    residuo_top = sorted(d['res_comp'], key=lambda rc: -(rc.get('contatos') or 0))[:2]
    res_str = ', '.join(
        f"{rc['padrao']} ({pct(rc.get('%', 0), 1)})"
        for rc in residuo_top
    )

    table_rows = [
        ('Período analisado', sem_display),
        ('Fonte dos dados', 'Treble — atendimento-treble e log de sessões'),
        ('Volume do período', f'{conversas} conversas atribuídas'),
        ('Responsável', 'Nexforce · RevOps'),
        ('% de conversas atendidas',
         f'{pct(pct_24h, 1)} ({resp_24h} de {conversas})'),
        ('Tempo mediano de 1ª resposta', f'{num_fmt(tempo, 2)} h'),
        ('Maior % de resposta — vendedores', top_vendors),
        ('Padrões de resíduo mais frequentes', res_str),
    ]

    return f"""
<section class="cover">
  <div class="eyebrow">N E X F O R C E &nbsp; · &nbsp; R E V O P S</div>
  <h1>Relatório de Atendimento<br><span>Treble</span> · {CLIENTE}</h1>
  <p class="subtitle">Treble &nbsp;·&nbsp; {CLIENTE} &nbsp;·&nbsp; Nexforce RevOps</p>
  <div class="meta-row">
    <div class="meta-item">
      <div class="label">Período</div>
      <div class="val">{sem_display}</div>
    </div>
    <div class="meta-item">
      <div class="label">Fonte</div>
      <div class="val">Treble · atendimento-treble</div>
    </div>
    <div class="meta-item">
      <div class="label">Volume</div>
      <div class="val">{conversas} conversas</div>
    </div>
    <div class="meta-item">
      <div class="label">Responsável</div>
      <div class="val">Nexforce · RevOps</div>
    </div>
  </div>

  <div style="margin-top:48px">
    <div class="section-index">0 0 &nbsp; — &nbsp; S U M Á R I O &nbsp; E X E C U T I V O</div>
    <h2 class="section-title" style="font-size:22px;margin-bottom:20px">Visão geral do período</h2>
    {cards}
    {summary_table_html(table_rows)}
  </div>
</section>
"""


# ═══════════════════════════════════════════════════════════════════════
# SECTION 01 — CONVERSAS
# ═══════════════════════════════════════════════════════════════════════

def build_section_01(d):
    r     = d['resumo']
    total = int(r.get('Conversas atendidas (total)', 0) or 0)
    resp  = int(r.get('Conversas respondidas por vendedor', 0) or 0)
    sem   = int(r.get('Conversas SEM resposta do vendedor', 0) or 0)
    pct_s = float(r.get('% sem resposta', 0) or 0)
    sem_ini, sem_fim = semana_datas()

    return f"""
<section class="section">
{section_header_html('01', 'CONVERSAS',
    'Conversas atribuídas a um vendedor no período, separadas entre as que tiveram resposta registrada e as que não tiveram.')}
{caption_html('Conversas respondidas x sem resposta do vendedor (base: atendimento-treble).')}
{chart_tag('A1_resposta.png', 'Conversas respondidas x sem resposta')}
{callout_html(f'Das {total} conversas atribuídas a vendedores entre {sem_ini} e {sem_fim}, {resp} tiveram resposta registrada e {sem} não tiveram, o que corresponde a {pct(pct_s, 1)} sem resposta no período.')}
</section>
"""


# ═══════════════════════════════════════════════════════════════════════
# SECTION 02 — CONVERSAS ATENDIDAS (dois gráficos, cadência semanal)
# ═══════════════════════════════════════════════════════════════════════

def _dia_fmt(v):
    try:
        return v.strftime('%d/%m')
    except AttributeError:
        return str(v)


def build_section_02(d):
    r       = d['resumo']
    total   = int(r.get('Conversas atendidas (total)', 0) or 0)
    em_24h  = int(r.get('Respondidas dentro de 24h', 0) or 0)
    pct_24h = float(r.get('% respondidas em 24h (do total)', 0) or 0)

    # PADRÃO desde 06/07/2026: relatório SEMANAL usa cadência DIÁRIA (seg–sex,
    # o fluxo não recebe novos leads no fim de semana); relatório de PERÍODO
    # mantém a cadência semanal.
    if TIPO == 'semanal':
        return _section_02_diaria(d, total, em_24h, pct_24h)

    semanal = d.get('semanal_24h', [])
    if len(semanal) == 1:
        # semana única: melhor/pior/maior volume seriam a mesma barra — não há cruzamento
        unica = semanal[0]
        cheg = int(unica.get('chegaram') or 0)
        leitura = (
            f'No total do período, {em_24h} das {total} conversas foram atendidas ({pct(pct_24h, 1)}). '
            f'Os dois gráficos mostram a mesma semana: das {cheg} conversas que chegaram a um vendedor, '
            f'{em_24h} tiveram a primeira resposta dentro da janela de 24h.'
        )
    elif semanal:
        max_vol  = max(semanal, key=lambda x: x.get('chegaram') or 0)
        best_sem = max(semanal, key=lambda x: (x.get('respondidos_24h') or 0) / max(1, x.get('chegaram') or 1))
        worst_sem= min(semanal, key=lambda x: (x.get('respondidos_24h') or 0) / max(1, x.get('chegaram') or 1))
        max_chegaram = int(max_vol.get('chegaram') or 0)
        max_pct   = round(100 * (max_vol.get('respondidos_24h') or 0) / max(1, max_chegaram or 1), 1)
        best_pct  = round(100 * (best_sem.get('respondidos_24h') or 0) / max(1, best_sem.get('chegaram') or 1), 1)
        worst_pct = round(100 * (worst_sem.get('respondidos_24h') or 0) / max(1, worst_sem.get('chegaram') or 1), 1)
        worst_cheg = int(worst_sem.get('chegaram') or 0)
        best_cheg  = int(best_sem.get('chegaram') or 0)
        max_dt = max_vol.get('dia')
        max_str = max_dt.strftime('%d/%m') if hasattr(max_dt, 'strftime') else str(max_dt)
        best_dt = best_sem.get('dia')
        best_str = best_dt.strftime('%d/%m') if hasattr(best_dt, 'strftime') else str(best_dt)
        worst_dt = worst_sem.get('dia')
        worst_str = worst_dt.strftime('%d/%m') if hasattr(worst_dt, 'strftime') else str(worst_dt)
        # leitura cruzando os dois gráficos: volume absoluto (1º) x proporção (2º)
        leitura = (
            f'No total do período, {em_24h} das {total} conversas foram atendidas ({pct(pct_24h, 1)}). '
            f'Cruzando os dois gráficos: a semana de maior volume (iniciando em {max_str}, {max_chegaram} chegadas) '
            f'atendeu {pct(max_pct, 1)} delas, enquanto o melhor índice proporcional ocorreu na semana de {best_str} '
            f'({pct(best_pct, 1)} de {best_cheg} chegadas) e o menor na de {worst_str} ({pct(worst_pct, 1)} de {worst_cheg} chegadas).'
        )
    else:
        leitura = f'No total do período, {em_24h} das {total} conversas foram atendidas ({pct(pct_24h, 1)}).'

    return f"""
<section class="section">
{section_header_html('02', 'CONVERSAS ATENDIDAS',
    'Conversas que chegaram a um vendedor e quantas foram atendidas (primeira resposta em até 24h, janela da Meta), agrupadas por semana (domingo a sábado).')}

{caption_html('Chegada ao vendedor (barra azul) e conversas atendidas (barra verde, contida na azul), por semana — valores absolutos.')}
{chart_tag('B_24h.png', 'Chegada ao vendedor x conversas atendidas — semanal')}

{caption_html('Distribuição proporcional: fatia verde = conversas atendidas, fatia vermelha = não atendidas, cada barra soma 100% das chegadas daquela semana.')}
{chart_tag('B_24h_stacked.png', 'Conversas atendidas — distribuição proporcional por semana')}

{callout_html(leitura)}
</section>
"""


def _section_02_diaria(d, total, em_24h, pct_24h):
    """S.02 em cadência DIÁRIA (seg–sex) — padrão do relatório semanal."""
    diario = [x for x in d.get('diario_24h', []) if (x.get('chegaram') or 0) > 0]
    if len(diario) >= 2:
        taxa = lambda x: (x.get('respondidos_24h') or 0) / max(1, x.get('chegaram') or 1)
        max_vol = max(diario, key=lambda x: x.get('chegaram') or 0)
        melhor  = max(diario, key=taxa)
        pior    = min(diario, key=taxa)
        leitura = (
            f'No total do período, {em_24h} das {total} conversas foram atendidas ({pct(pct_24h, 1)}). '
            f'Cruzando os dois gráficos: o dia de maior volume ({_dia_fmt(max_vol.get("dia"))}, '
            f'{int(max_vol.get("chegaram") or 0)} chegadas) atendeu {pct(100 * taxa(max_vol), 1)} delas; '
            f'o melhor índice proporcional ocorreu em {_dia_fmt(melhor.get("dia"))} '
            f'({pct(100 * taxa(melhor), 1)} de {int(melhor.get("chegaram") or 0)} chegadas) '
            f'e o menor em {_dia_fmt(pior.get("dia"))} '
            f'({pct(100 * taxa(pior), 1)} de {int(pior.get("chegaram") or 0)} chegadas).'
        )
    elif len(diario) == 1:
        unico = diario[0]
        leitura = (
            f'No total do período, {em_24h} das {total} conversas foram atendidas ({pct(pct_24h, 1)}). '
            f'Todas as chegadas se concentraram em {_dia_fmt(unico.get("dia"))} '
            f'({int(unico.get("chegaram") or 0)} conversas).'
        )
    else:
        leitura = f'No total do período, {em_24h} das {total} conversas foram atendidas ({pct(pct_24h, 1)}).'

    return f"""
<section class="section">
{section_header_html('02', 'CONVERSAS ATENDIDAS',
    'Conversas que chegaram a um vendedor e quantas foram atendidas (primeira resposta em até 24h, janela da Meta), agrupadas por dia útil.')}

{caption_html('Chegada ao vendedor (barra azul) e conversas atendidas (barra verde, contida na azul), por dia — valores absolutos.')}
{chart_tag('B_24h_dia.png', 'Chegada ao vendedor x conversas atendidas — diário')}

{caption_html('Distribuição proporcional: fatia verde = conversas atendidas, fatia vermelha = não atendidas, cada barra soma 100% das chegadas daquele dia.')}
{chart_tag('B_24h_dia_stacked.png', 'Conversas atendidas — distribuição proporcional por dia')}

{callout_html(leitura)}
</section>
"""


# ═══════════════════════════════════════════════════════════════════════
# SECTION 03 — VOLUME POR VENDEDOR
# ═══════════════════════════════════════════════════════════════════════

def build_section_03(d):
    vendedores = d['vendedores']
    total_chats = sum(int(v.get('chats_recebidos') or 0) for v in vendedores)
    n_vend = len(vendedores)

    # ranking de % de resposta só entre quem recebeu conversa no período;
    # vendedor com 0 recebidas teria 100% trivial e distorceria a leitura
    com_recebidas = [v for v in vendedores if (v.get('chats_recebidos') or 0) > 0]

    if com_recebidas:
        pct_de = lambda v: round(100 - (v.get('%_nao_respondidos') or 0), 1)
        vol_de = lambda v: int(v.get('chats_recebidos') or 0)
        mais_chats = max(com_recebidas, key=vol_de)
        pct_top_vol = pct_de(mais_chats)

        # desempate EXPLÍCITO (regra de 06/07/2026): entre vendedores empatados
        # no mesmo percentual, prevalece o de maior volume recebido — e a
        # leitura declara o critério quando há empate.
        max_pct = max(pct_de(v) for v in com_recebidas)
        empat_max = [v for v in com_recebidas if pct_de(v) == max_pct]
        mais_resp = max(empat_max, key=vol_de)
        min_pct = min(pct_de(v) for v in com_recebidas)
        empat_min = [v for v in com_recebidas if pct_de(v) == min_pct]
        menos_resp = max(empat_min, key=vol_de)

        def _nota_desempate(empatados, escolhido, p):
            # desempate por maior volume, dito de forma SUTIL (regra de 06/07/2026):
            # explica o porquê sem anunciar "critério de desempate"
            if len(empatados) < 2:
                return ''
            return (f', por ter o maior volume ({vol_de(escolhido)} conversas) entre os '
                    f'vendedores com {pct(p, 0)} de resposta')

        leitura = (
            f'{len(com_recebidas)} vendedores receberam ao todo {total_chats} conversas no período. '
            f'{mais_chats.get("vendedor")} liderou em volume ({vol_de(mais_chats)} conversas) '
            f'com {pct(pct_top_vol, 0)} de resposta. '
            f'{mais_resp.get("vendedor")} teve o maior percentual de resposta '
            f'({pct(max_pct, 0)}){_nota_desempate(empat_max, mais_resp, max_pct)}, '
            f'e {menos_resp.get("vendedor")} o menor '
            f'({pct(min_pct, 0)}){_nota_desempate(empat_min, menos_resp, min_pct)}.'
        )
    else:
        leitura = 'Nenhum vendedor com dados no período.'

    return f"""
<section class="section">
{section_header_html('03', 'VOLUME POR VENDEDOR',
    'Volume de conversas recebidas por cada vendedor no período e o percentual de conversas respondidas, calculado sobre o total recebido por cada um.')}
{caption_html('Conversas recebidas por vendedor.')}
{chart_tag('D1_recebidos.png', 'Conversas recebidas por vendedor')}
{caption_html('% de conversas respondidas por vendedor.')}
{chart_tag('D2_resp.png', '% de conversas respondidas por vendedor')}
{callout_html(leitura)}
</section>
"""


# ═══════════════════════════════════════════════════════════════════════
# SECTION 04 — TEMPO DE RESPOSTA
# ═══════════════════════════════════════════════════════════════════════

def build_section_04(d):
    respondentes = [v for v in d['vendedores'] if (v.get('respondidos') or 0) > 0]
    respondentes_s = sorted(respondentes, key=lambda v: v.get('tempo_med_1a_resp_h') or 99)
    tempo_semanal = d.get('d3_tempo_semanal', [])

    # o gráfico de evolução (D3) é linha única do time; eixo y em h ou min (--d3-unidade)
    unid_lbl = 'minutos' if D3_UNIDADE == 'min' else 'horas'

    def tempo_d3_fmt(horas):
        if D3_UNIDADE == 'min':
            return f'{num_fmt(float(horas) * 60, 0)} min'
        return f'{num_fmt(horas, 2)} h'

    def dia_fmt(v):
        try:
            return v.strftime('%d/%m')
        except AttributeError:
            return str(v)

    if respondentes_s:
        mais_rapido = respondentes_s[0]
        mais_lento  = respondentes_s[-1]
        semanas = [r for r in tempo_semanal if r.get('tempo_med_h') is not None]
        if semanas:
            melhor = min(semanas, key=lambda r: float(r['tempo_med_h']))
            pior   = max(semanas, key=lambda r: float(r['tempo_med_h']))
            leitura = (
                f'O tempo mediano de 1ª resposta do time variou de '
                f'{tempo_d3_fmt(melhor["tempo_med_h"])} (semana de {dia_fmt(melhor["dia"])}) a '
                f'{tempo_d3_fmt(pior["tempo_med_h"])} (semana de {dia_fmt(pior["dia"])}). '
                f'No acumulado do período, {mais_rapido.get("vendedor")} teve o menor tempo mediano '
                f'({num_fmt(mais_rapido.get("tempo_med_1a_resp_h"), 2)} h), '
                f'e {mais_lento.get("vendedor")} o maior ({num_fmt(mais_lento.get("tempo_med_1a_resp_h"), 2)} h).'
            )
        else:
            leitura = (
                f'{len(respondentes)} vendedores registraram resposta no período. '
                f'Os tempos medianos variaram de '
                f'{num_fmt(mais_rapido.get("tempo_med_1a_resp_h"), 2)} h ({mais_rapido.get("vendedor")}) '
                f'a {num_fmt(mais_lento.get("tempo_med_1a_resp_h"), 2)} h ({mais_lento.get("vendedor")}).'
            )
    else:
        leitura = 'Nenhum vendedor registrou resposta no período.'

    chart_temporal = ''
    if tempo_semanal:
        arquivo_d3 = 'D3_tempo_semanal_min.png' if D3_UNIDADE == 'min' else 'D3_tempo_semanal.png'
        desc_04 = ('Tempo mediano entre a atribuição e a primeira mensagem do vendedor. '
                   f'O primeiro gráfico mostra a evolução semanal do time (em {unid_lbl}); '
                   'o segundo, o resumo do período por vendedor (em horas).')
        chart_temporal = (
            caption_html(f'Tempo mediano de 1ª resposta do time ao longo das semanas ({unid_lbl}).')
            + '\n' + chart_tag(arquivo_d3, 'Evolução semanal do tempo mediano de resposta do time')
            + '\n'
        )
    else:
        desc_04 = 'Tempo mediano, em horas, entre a atribuição e a primeira mensagem do vendedor.'

    return f"""
<section class="section">
{section_header_html('04', 'TEMPO DE RESPOSTA', desc_04)}
{chart_temporal}{caption_html('Tempo mediano de 1ª resposta por vendedor — resumo do período (horas).')}
{chart_tag('D4_tempo.png', 'Tempo mediano de 1ª resposta — resumo do período')}
{callout_html(leitura)}
</section>
"""


# ═══════════════════════════════════════════════════════════════════════
# SECTION 05 — PADRÃO DE RESÍDUO
# ═══════════════════════════════════════════════════════════════════════

def build_section_05(d):
    res_comp = d['res_comp']
    res_resp = d['res_resp']

    hosp_rows_comp = [rc for rc in res_comp if 'Hospitalar' in str(rc.get('padrao', ''))]
    ind_rows_comp  = [rc for rc in res_comp if 'Industrial' in str(rc.get('padrao', ''))]
    total_n   = sum(rc.get('contatos') or 0 for rc in res_comp)
    hosp_n    = sum(rc.get('contatos') or 0 for rc in hosp_rows_comp)
    hosp_pct_sum = sum(rc.get('%') or 0 for rc in hosp_rows_comp)
    ind_n     = sum(rc.get('contatos') or 0 for rc in ind_rows_comp)
    ind_pct   = round(100 * ind_n / total_n, 1) if total_n else 0

    hosp_resp_rows = [rc for rc in res_resp if 'Hospitalar' in str(rc.get('padrao', ''))]
    ind_resp_rows  = [rc for rc in res_resp if 'Industrial' in str(rc.get('padrao', ''))]
    hosp_total_c   = sum(rc.get('contatos') or 0 for rc in hosp_resp_rows)
    hosp_sem_c     = sum(rc.get('sem_resposta') or 0 for rc in hosp_resp_rows)
    hosp_sem_pct   = round(100 * hosp_sem_c / hosp_total_c, 1) if hosp_total_c else 0
    ind_sem_pct    = 100.0 if ind_n > 0 and all((rc.get('%_sem_resposta') or 0) == 100 for rc in ind_resp_rows) else None

    ind_note = (f'Os {ind_n} contatos do padrão Industrial ({pct(ind_pct, 1)} do total) '
                f'ficaram 100% sem resposta — nenhum chegou a ser atendido. ') if ind_sem_pct == 100 else ''

    leitura = (
        f'O padrão Hospitalar concentra a maioria dos contatos: {hosp_n} de {total_n} ({pct(hosp_pct_sum, 1)}), '
        f'com {pct(hosp_sem_pct, 1)} sem resposta. '
        + ind_note +
        f'A tabela acima indica o tempo mediano de resposta para cada padrão, '
        f'permitindo comparar não só quem foi atendido, mas também com que agilidade.'
    )

    return f"""
<section class="section">
{section_header_html('05', 'PADRÃO DE RESÍDUO',
    'Composição dos contatos que chegaram a um vendedor por padrão de resíduo informado no fluxo, e taxa de resposta por padrão.')}
{caption_html('Composição dos contatos por padrão de resíduo.')}
{chart_tag('R1_residuo_comp.png', 'Composição por padrão de resíduo')}
{caption_html('Conversas respondidas x sem resposta por padrão de resíduo (volume, % dentro de cada barra).')}
{chart_tag('R2_residuo_resp.png', 'Respondidos x sem resposta por padrão')}
{residuo_table_html(res_resp)}
{callout_html(leitura)}
</section>
"""


# ═══════════════════════════════════════════════════════════════════════
# SECTION 06 — QUALIFICAÇÃO
# ═══════════════════════════════════════════════════════════════════════

# ordem real dos campos em cada caminho (a mesma exibida nos gráficos G2b/G2c;
# desde 06/07/2026 as abas G2b/G2c também saem nesta ordem — ver _analise_comum)
_FLOW_NOVO = ['email', 'nome', 'empresa', 'cpf/cnpj', 'estado', 'cidade',
              'classe_residuo', 'tipo_residuo', 'qtd_residuo', 'tipo_servico']
_FLOW_CLI  = ['email', 'nome', 'cpf/cnpj', 'tipo_residuo', 'qtd_residuo', 'tipo_servico']
_NOME_CAMPO = {'email': 'e-mail', 'cpf/cnpj': 'CPF/CNPJ',
               'classe_residuo': 'classe do resíduo', 'tipo_residuo': 'tipo de resíduo',
               'qtd_residuo': 'quantidade de resíduo', 'tipo_servico': 'tipo de serviço'}

_nome_campo = lambda c: _NOME_CAMPO.get(c, c)


def _quedas_fluxo(comp, flow_order):
    """Frase de quedas de um fluxo (queda total + 2 maiores quedas consecutivas).

    Retorna (frase, pct por campo, ordem real, peso das quedas por campo),
    onde o peso soma as quedas das 2 maiores em que cada campo participa —
    usado para escolher o campo citado no cruzamento entre os fluxos.
    """
    pct_campo = {str(r.get('campo', '')).strip().lower(): float(r.get('%_preenchido') or 0)
                 for r in comp}
    ordem = [c for c in flow_order if c in pct_campo]
    if len(ordem) < 3:
        return '', pct_campo, ordem, {}
    quedas = [(ordem[i], ordem[i + 1], pct_campo[ordem[i]] - pct_campo[ordem[i + 1]])
              for i in range(len(ordem) - 1)]
    # só quedas REAIS (> 0): nunca citar um degrau de 0,0 pp como "maior queda"
    maiores = [q for q in sorted(quedas, key=lambda q: -q[2]) if q[2] > 0][:2]
    peso_campo = {}
    for a, b, delta in maiores:
        for c in (a, b):
            peso_campo[c] = peso_campo.get(c, 0) + delta
    txt = (
        f' Do primeiro campo ({_nome_campo(ordem[0])}, {pct(pct_campo[ordem[0]], 1)}) ao último '
        f'({_nome_campo(ordem[-1])}, {pct(pct_campo[ordem[-1]], 1)}), o preenchimento cai '
        f'{num_fmt(pct_campo[ordem[0]] - pct_campo[ordem[-1]], 1)} pontos percentuais'
    )
    if len(maiores) >= 2:
        txt += (
            f'; as maiores quedas entre campos consecutivos ocorrem de {_nome_campo(maiores[0][0])} '
            f'({pct(pct_campo[maiores[0][0]], 1)}) para {_nome_campo(maiores[0][1])} '
            f'({pct(pct_campo[maiores[0][1]], 1)}) e de {_nome_campo(maiores[1][0])} '
            f'({pct(pct_campo[maiores[1][0]], 1)}) para {_nome_campo(maiores[1][1])} '
            f'({pct(pct_campo[maiores[1][1]], 1)}).'
        )
    elif len(maiores) == 1:
        txt += (
            f'; a queda concentra-se em um único degrau, de {_nome_campo(maiores[0][0])} '
            f'({pct(pct_campo[maiores[0][0]], 1)}) para {_nome_campo(maiores[0][1])} '
            f'({pct(pct_campo[maiores[0][1]], 1)}).'
        )
    else:
        txt += '.'
    return txt, pct_campo, ordem, peso_campo


def _cruzamento_fluxos(pct_novo, ordem_novo, peso_novo, pct_cli, ordem_cli, peso_cli):
    """Frase que relaciona os dois gráficos da seção 06, sempre calculada dos dados."""
    if not ordem_novo or not ordem_cli:
        return ''
    campo1 = ordem_cli[0]
    v_cli  = pct_cli[campo1]
    v_novo = pct_novo.get(ordem_novo[0], 0)
    patamar = 'mais baixo' if v_cli < v_novo else 'mais alto'
    txt = (
        f' Na comparação entre os caminhos, o "Já sou cliente" parte de um patamar {patamar} '
        f'desde o primeiro campo ({_nome_campo(campo1)}: {pct(v_cli, 1)}, contra '
        f'{pct(v_novo, 1)} no "Sou novo por aqui")'
    )
    comuns = set(peso_novo) & set(peso_cli)
    if comuns:
        destaque = max(comuns, key=lambda c: peso_novo[c] + peso_cli[c])
        txt += (f', e o {_nome_campo(destaque)} aparece entre as maiores quedas '
                'consecutivas nos dois fluxos.')
    else:
        txt += '.'
    return txt


def build_section_06(d):
    comp_novo = d['comp_novo']
    comp_cli  = d['comp_cli']
    caminhos  = d['caminhos']

    novo_n = next((int(c.get('conversas') or 0) for c in caminhos
                   if 'novo' in str(c.get('caminho', '')).lower()), 0)
    cli_n  = next((int(c.get('conversas') or 0) for c in caminhos
                   if 'cliente' in str(c.get('caminho', '')).lower()), 0)

    quedas_novo, pct_novo, ordem_novo, peso_novo = _quedas_fluxo(comp_novo, _FLOW_NOVO)
    quedas_cli,  pct_cli,  ordem_cli,  peso_cli  = _quedas_fluxo(comp_cli, _FLOW_CLI)

    # cada gráfico tem a própria Leitura; o cruzamento entre os fluxos fecha a segunda
    leitura_novo = (
        f'No fluxo "Sou novo por aqui", {novo_n} sessões escolheram esse caminho. '
        'A queda de campo a campo reflete o abandono ao longo do preenchimento.'
        + quedas_novo
    )
    leitura_cli = (
        f'No fluxo "Já sou cliente", {cli_n} sessões escolheram esse caminho.'
        + quedas_cli
        + _cruzamento_fluxos(pct_novo, ordem_novo, peso_novo, pct_cli, ordem_cli, peso_cli)
    )

    return f"""
<section class="section">
{section_header_html('06', 'QUALIFICAÇÃO',
    'Percentual de preenchimento de cada campo, calculado sobre as sessões que entraram em cada caminho (a partir da primeira pergunta do fluxo). Sessões sem caminho definido ficam fora desta análise.')}
{caption_html(f'Preenchimento dos campos — Sou novo por aqui ({novo_n} sessões neste caminho).')}
{chart_tag('G2b_novo.png', 'Preenchimento campos — Sou novo por aqui')}
{callout_html(leitura_novo)}
{caption_html(f'Preenchimento dos campos — Já sou cliente ({cli_n} sessões neste caminho).')}
{chart_tag('G2c_cli.png', 'Preenchimento campos — Já sou cliente')}
{callout_html(leitura_cli)}
</section>
"""


# ═══════════════════════════════════════════════════════════════════════
# CLOSING
# ═══════════════════════════════════════════════════════════════════════

def build_closing(semana_str):
    sem_ini, sem_fim = semana_datas()
    return f"""
<div class="closing">
  <div class="eyebrow">N E X F O R C E &nbsp; · &nbsp; R E V O P S</div>
  <p style="color:var(--muted);margin-bottom:8px">Uso interno.</p>
  <p>Relatório gerado pela Nexforce (RevOps) a partir dos dados operacionais exportados da Treble, cobrindo a semana de {sem_ini} a {sem_fim}. Indicadores calculados sobre a base atendimento-treble e o log de sessões.</p>
  <div style="margin-top:24px;font-family:'JetBrains Mono','Courier New',monospace;font-size:10px;color:var(--muted)">
    Nexforce · RevOps &nbsp;·&nbsp; {semana_str.replace('-', '/')}
  </div>
</div>
"""


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def _semana_anterior_fallback():
    from datetime import timedelta
    today = datetime.today()
    weekday = today.weekday()
    days_to_sat = (weekday + 2) % 7 or 7
    last_sat = today - timedelta(days=days_to_sat)
    last_sun = last_sat - timedelta(days=6)
    return f'{last_sun.strftime("%d-%m-%Y")} a {last_sat.strftime("%d-%m-%Y")}'


def main():
    global SEMANA, _semana_cache
    _semana_cache = None  # reseta cache ao rodar main

    if SEMANA is None:
        import re as _re
        try:
            wb = openpyxl.load_workbook(XLSX_ANALISE, read_only=True)
            found = None
            for sname in wb.sheetnames:
                if found:
                    break
                ws = wb[sname]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            m = _re.search(r'(\d{2}-\d{2}-\d{4})\s+a\s+(\d{2}-\d{2}-\d{4})',
                                           str(cell))
                            if m:
                                found = f'{m.group(1)} a {m.group(2)}'
                                break
                    if found:
                        break
            wb.close()
            SEMANA = found
        except Exception:
            pass
        if SEMANA is None:
            if TIPO == 'periodo':
                print('ERRO: Período não detectado no xlsx.')
                import sys; sys.exit(1)
            SEMANA = _semana_anterior_fallback()
            print(f'[AVISO] Período não detectado no xlsx — usando semana anterior: {SEMANA}')

    if TIPO == 'semanal':
        output = f'Relatório semanal Treble - {SEMANA}.html'
    else:
        output = f'Relatório Treble - Utilização geral - {SEMANA}.html'

    semana_str = SEMANA
    print(f'Tipo: {TIPO} | Período: {semana_str}')
    print(f'Carregando dados de {XLSX_ANALISE}...')
    d = load_data()

    print('Gerando HTML...')

    body = (
        build_cover(d, semana_str)
        + build_section_01(d)
        + build_section_02(d)
        + build_section_03(d)
        + build_section_04(d)
        + build_section_05(d)
        + build_section_06(d)
        + build_closing(semana_str)
    )

    # Google Fonts com fallback local — falha silenciosa se offline
    google_fonts = (
        '<link rel="preconnect" href="https://fonts.googleapis.com">'
        '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>'
        '<link href="https://fonts.googleapis.com/css2?'
        'family=JetBrains+Mono:wght@400;600;700'
        '&family=Lato:wght@400;700&display=swap" rel="stylesheet">'
    )

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Relatório de Atendimento Treble · {CLIENTE} — {semana_str}</title>
{google_fonts}
<style>
{CSS}
</style>
</head>
<body>
<div class="doc">
{body}
</div>
</body>
</html>"""

    with open(output, 'w', encoding='utf-8') as f:
        f.write(html)

    size = os.path.getsize(output)
    print(f'Relatório gerado: {output} ({size:,} bytes)')
    return output


if __name__ == '__main__':
    import argparse as _ap
    _parser = _ap.ArgumentParser(description='Gera o relatório Treble em HTML.')
    _parser.add_argument('--tipo', choices=['semanal', 'periodo'], default=None)
    _parser.add_argument('--xlsx',     default=None)
    _parser.add_argument('--graficos', default=None)
    _parser.add_argument('--semana',   default=None)
    _parser.add_argument('--d3-unidade', choices=['h', 'min'], default='h',
                         dest='d3_unidade',
                         help='unidade do eixo y da evolução semanal (S.04); padrão aprovado: h')
    _parser.add_argument('--com-titulos', action='store_true', dest='com_titulos',
                         help='inclui os títulos cinza (captions) acima dos gráficos; '
                              'o padrão aprovado é sem títulos')
    _args = _parser.parse_args()

    D3_UNIDADE  = _args.d3_unidade
    SEM_TITULOS = not _args.com_titulos

    if _args.tipo == 'periodo':
        TIPO         = 'periodo'
        XLSX_ANALISE = _args.xlsx     or 'relatorio_equipe_periodo.xlsx'
        GRAFICOS     = _args.graficos or 'graficos/equipe'
    elif _args.tipo == 'semanal':
        TIPO         = 'semanal'
        XLSX_ANALISE = _args.xlsx     or XLSX_ANALISE
        GRAFICOS     = _args.graficos or GRAFICOS
    else:
        if _args.xlsx:     XLSX_ANALISE = _args.xlsx
        if _args.graficos: GRAFICOS     = _args.graficos

    if _args.semana:
        SEMANA = _args.semana

    main()
