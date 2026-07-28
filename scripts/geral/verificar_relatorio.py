#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
verificar_relatorio.py
Verifica a integridade do relatório Treble antes da entrega.

Uso:
  python scripts/geral/verificar_relatorio.py
  python scripts/geral/verificar_relatorio.py --tipo periodo
  python scripts/geral/verificar_relatorio.py --html "Relatório semanal Treble - 07-06-2026 a 13-06-2026.html"
  python scripts/geral/verificar_relatorio.py --tipo periodo --html "Relatório Treble - Utilização geral - 01-05-2026 a 31-05-2026.html"

Rodar sempre da RAIZ do projeto.
"""

import os
import sys
import re
import argparse
from datetime import datetime

DATE_PATTERN = re.compile(r'(\d{2}-\d{2}-\d{4})\s+a\s+(\d{2}-\d{2}-\d{4})')

SHEETS_REQUIRED = [
    'A1_resumo', 'D1_por_vendedor', 'F0_resumo_sessoes',
    'G1_caminhos', 'G2b_completude_novo', 'G2c_completude_cli',
    'B_24h_diario', 'B_24h_semanal', 'R1_residuo_comp', 'R2_residuo_resp',
]

# S.02 (regra de 06/07/2026): semanal usa cadência DIÁRIA (seg–sex);
# periodo mantém cadência semanal
PNGS_COMUNS = {
    'A1_resposta.png':    'S.01 — Conversas respondidas x sem resposta',
    'D1_recebidos.png':   'S.03a — Conversas recebidas por vendedor',
    'D2_resp.png':        'S.03b — % de conversas respondidas por vendedor',
    'D4_tempo.png':       'S.04 — Tempo mediano de 1ª resposta',
    'R1_residuo_comp.png':'S.05a — Composição por padrão de resíduo',
    'R2_residuo_resp.png':'S.05b — Respondidos x sem resposta por padrão',
    'G2b_novo.png':       'S.06a — Completude — Sou novo por aqui',
    'G2c_cli.png':        'S.06b — Completude — Já sou cliente',
}
PNGS_POR_TIPO = {
    'semanal': {
        'B_24h_dia.png':         'S.02a — Chegada x conversas atendidas (diário seg–sex, absoluto)',
        'B_24h_dia_stacked.png': 'S.02b — Conversas atendidas — proporcional por dia (empilhado)',
        **PNGS_COMUNS,
    },
    'periodo': {
        'B_24h.png':         'S.02a — Chegada x conversas atendidas (semanal, absoluto)',
        'B_24h_stacked.png': 'S.02b — Conversas atendidas — proporcional por semana (empilhado)',
        **PNGS_COMUNS,
    },
}

DEFAULTS = {
    'semanal': ('relatorio_treble_semanal.xlsx', 'graficos/treble'),
    'periodo': ('relatorio_equipe_periodo.xlsx', 'graficos/equipe'),
}


class Checker:
    def __init__(self):
        self.items = []  # (section, status, msg)
        self._current_section = 'Geral'

    def section(self, title):
        self._current_section = title

    def ok(self, msg):
        self.items.append((self._current_section, 'ok', msg))

    def fail(self, msg):
        self.items.append((self._current_section, 'fail', msg))

    def warn(self, msg):
        self.items.append((self._current_section, 'warn', msg))

    @property
    def n_fail(self):
        return sum(1 for _, s, _ in self.items if s == 'fail')

    @property
    def n_ok(self):
        return sum(1 for _, s, _ in self.items if s == 'ok')

    @property
    def n_warn(self):
        return sum(1 for _, s, _ in self.items if s == 'warn')

    def check_xlsx(self, xlsx_path, tipo):
        self.section('Dados (xlsx)')

        if not os.path.exists(xlsx_path):
            self.fail(f'xlsx não encontrado: {xlsx_path}')
            return None

        size_kb = os.path.getsize(xlsx_path) / 1024
        self.ok(f'xlsx encontrado: {os.path.basename(xlsx_path)} ({size_kb:.0f} KB)')

        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            sheet_names = wb.sheetnames
        except ImportError:
            self.fail('openpyxl não instalado — pip install openpyxl')
            return None
        except Exception as e:
            self.fail(f'Erro ao abrir xlsx: {e}')
            return None

        # Sheets
        missing = [s for s in SHEETS_REQUIRED if s not in sheet_names]
        self.ok(f'{len(SHEETS_REQUIRED) - len(missing)}/{len(SHEETS_REQUIRED)} abas obrigatórias presentes')
        for s in missing:
            self.fail(f'Aba obrigatória ausente: {s}')

        # Read A1_resumo
        resumo = {}
        if 'A1_resumo' in sheet_names:
            ws = wb['A1_resumo']
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if row[0] is not None:
                    resumo[str(row[0])] = row[1]

        # Total de conversas
        total_raw = resumo.get('Conversas atendidas (total)')
        if total_raw is None:
            self.warn('Métrica "Conversas atendidas (total)" não encontrada em A1_resumo')
            total = None
        else:
            total = int(total_raw)
            if total == 0:
                self.fail('Total de conversas é 0 — dados podem estar vazios ou filtro errado')
            else:
                self.ok(f'Total de conversas: {total}')

        # Sanity: respondidas + sem resposta = total
        resp_raw = resumo.get('Conversas respondidas por vendedor')
        sem_raw  = resumo.get('Conversas SEM resposta do vendedor')
        if resp_raw is not None and sem_raw is not None and total is not None:
            resp, sem = int(resp_raw), int(sem_raw)
            if abs(resp + sem - total) <= 1:
                self.ok(f'Totais consistentes: {resp} respondidas + {sem} sem resposta = {total}')
            else:
                self.fail(
                    f'Totais inconsistentes: {resp} respondidas + {sem} sem resposta = {resp + sem}, '
                    f'mas total = {total}  (diferença: {abs(resp + sem - total)})'
                )
        else:
            self.warn('Não foi possível verificar consistência dos totais (campos ausentes em A1_resumo)')

        # Sanity: % sem resposta plausível
        pct_sem = resumo.get('% sem resposta')
        if pct_sem is not None:
            p = float(pct_sem)
            if 0.0 <= p <= 100.0:
                self.ok(f'% sem resposta: {p:.1f}% (valor plausível)')
            else:
                self.fail(f'% sem resposta fora do intervalo 0-100: {p}')

        # Sanity: tem vendedores com dados?
        if 'D1_por_vendedor' in sheet_names:
            ws = wb['D1_por_vendedor']
            rows = list(ws.iter_rows(values_only=True))
            data_rows = [r for r in rows[1:] if any(v is not None for v in r)]
            if len(data_rows) == 0:
                self.fail('D1_por_vendedor sem linhas de dados — base de vendedores vazia')
            else:
                self.ok(f'D1_por_vendedor: {len(data_rows)} vendedor(es) com dados')

        # Ordem das abas de completude por caminho = ordem REAL do fluxo
        # (regra permanente de 06/07/2026 — gráfico e Leitura de quedas na mesma ordem)
        try:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from _analise_comum import FLOW_NOVO, FLOW_CLI
            for aba, ordem in (('G2b_completude_novo', FLOW_NOVO),
                               ('G2c_completude_cli', FLOW_CLI)):
                if aba not in sheet_names:
                    continue
                ws = wb[aba]
                rows = list(ws.iter_rows(values_only=True))
                campos = [str(r[0]).strip().lower() for r in rows[1:]
                          if r and r[0] is not None]
                esperado = [c for c in ordem if c in campos]
                if campos == esperado:
                    self.ok(f'{aba}: campos na ordem real do fluxo')
                else:
                    self.fail(f'{aba}: campos fora da ordem do fluxo '
                              f'(atual: {campos} | esperado: {esperado})')
        except ImportError:
            self.warn('_analise_comum não importável — ordem G2b/G2c não verificada')

        # Balanço de transferências: soma out == soma in (semântica corrigida em 06/07/2026)
        if 'D1_por_vendedor' in sheet_names:
            ws = wb['D1_por_vendedor']
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(h) for h in rows[0]]
                if 'transferiu_out' in header and 'recebidos_via_transfer' in header:
                    i_out = header.index('transferiu_out')
                    i_in = header.index('recebidos_via_transfer')
                    soma_out = sum(int(r[i_out] or 0) for r in rows[1:] if r and r[0] is not None)
                    soma_in = sum(int(r[i_in] or 0) for r in rows[1:] if r and r[0] is not None)
                    if soma_out == soma_in:
                        self.ok(f'Transferências balanceadas: {soma_out} enviadas = {soma_in} recebidas')
                    else:
                        self.warn(f'Transferências desbalanceadas: {soma_out} enviadas x {soma_in} '
                                  'recebidas (pode ser legítimo se houver agente não canônico envolvido)')

        # Extract period
        period = None
        for sname in sheet_names:
            if period:
                break
            ws = wb[sname]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        m = DATE_PATTERN.search(str(cell))
                        if m:
                            period = f'{m.group(1)} a {m.group(2)}'
                            break
                if period:
                    break

        if period:
            self.ok(f'Período detectado no xlsx: {period}')
            try:
                d1 = datetime.strptime(period.split(' a ')[0], '%d-%m-%Y')
                d2 = datetime.strptime(period.split(' a ')[1], '%d-%m-%Y')
                days = (d2 - d1).days + 1
                if tipo == 'semanal':
                    if 6 <= days <= 8:
                        self.ok(f'Duração: {days} dias (semanal OK)')
                    else:
                        self.warn(f'Duração: {days} dias — esperado 7 para tipo semanal')
                else:
                    if days >= 28:
                        self.ok(f'Duração: {days} dias (período estendido OK)')
                    else:
                        self.warn(f'Duração: {days} dias — para tipo periodo, esperado > 28 dias')
            except Exception:
                pass
        else:
            self.warn('Período não detectado automaticamente no xlsx — verificar manualmente')

        wb.close()
        return period

    def check_pngs(self, graficos_path, tipo='semanal'):
        self.section('Gráficos (10 PNGs)')
        pngs_required = PNGS_POR_TIPO.get(tipo, PNGS_POR_TIPO['semanal'])

        if not os.path.isdir(graficos_path):
            self.fail(f'Pasta de gráficos não encontrada: {graficos_path}')
            for png in pngs_required:
                self.fail(f'PNG ausente (pasta inexistente): {png}')
            return

        self.ok(f'Pasta de gráficos encontrada: {graficos_path}')
        for png, label in pngs_required.items():
            path = os.path.join(graficos_path, png)
            if not os.path.exists(path):
                self.fail(f'Ausente: {png}  ({label})')
            elif os.path.getsize(path) == 0:
                self.fail(f'Vazio (0 bytes): {png}  ({label})')
            else:
                size_kb = os.path.getsize(path) / 1024
                self.ok(f'{png} ({size_kb:.0f} KB)')

    def check_html(self, html_path, tipo):
        self.section('HTML')

        if not os.path.exists(html_path):
            self.fail(f'HTML não encontrado: {html_path}')
            return

        size_kb = os.path.getsize(html_path) / 1024
        if size_kb < 50:
            self.fail(f'HTML suspeito — tamanho muito pequeno: {size_kb:.0f} KB (esperado >500 KB com imagens embutidas)')
        elif size_kb > 50000:
            self.warn(f'HTML com tamanho incomum: {size_kb:.0f} KB')
        else:
            self.ok(f'HTML encontrado: {os.path.basename(html_path)} ({size_kb:.0f} KB)')

        fname = os.path.basename(html_path)
        if DATE_PATTERN.search(fname):
            self.ok('Período detectado no nome do arquivo HTML')
        else:
            self.warn('Nome do HTML não contém padrão DD-MM-AAAA a DD-MM-AAAA')

        if not fname.lower().endswith('.html'):
            self.warn('Extensão do arquivo não é .html')

        if tipo == 'semanal' and 'semanal' not in fname.lower():
            self.warn('Tipo semanal mas "semanal" não está no nome do arquivo')
        if tipo == 'periodo' and 'semanal' in fname.lower():
            self.warn('Tipo periodo mas nome do arquivo contém "semanal" — verificar nomenclatura')

        # padrão canônico aprovado em 03/07/2026 (relatório 22/04–30/06)
        try:
            with open(html_path, encoding='utf-8') as f:
                html = f.read()
        except Exception as e:
            self.warn(f'Não foi possível ler o conteúdo do HTML para os checks de padrão: {e}')
            return

        n_captions = html.count('class="caption"')
        if n_captions:
            self.fail(f'{n_captions} título(s) cinza (caption) acima dos gráficos — '
                      'o padrão aprovado é SEM captions (não usar --com-titulos)')
        else:
            self.ok('Sem captions acima dos gráficos (padrão aprovado)')

        # o relatório é factual: não deve conter blocos de sugestão/iniciativa
        if 'Iniciativa:' in html:
            self.fail('Bloco "Iniciativa:" encontrado — o relatório não sugere ações')
        else:
            self.ok('Sem bloco "Iniciativa:" (padrão)')

        n_leituras_s06 = html.count('<strong>Leitura:</strong> No fluxo')
        if n_leituras_s06 == 2:
            self.ok('Seção 06 com uma Leitura por gráfico (2 blocos)')
        else:
            self.fail(f'Seção 06 com {n_leituras_s06} Leitura(s) iniciando com "No fluxo" — '
                      'o padrão é 2 (uma após cada gráfico)')

        # regra absoluta (06/07/2026): sem informação equivalente repetida
        if '% de atendimento realizado' in html:
            self.fail('"% de atendimento realizado" encontrado — métrica equivalente a '
                      '"% de conversas atendidas"; a regra é não repetir informação equivalente')
        else:
            self.ok('Sem "% de atendimento realizado" (regra de não repetição)')

        if '<th>% sem resposta</th>' in html:
            self.fail('Coluna "% sem resposta" na tabela de resíduo — é o complemento de '
                      '"% respondida"; a regra é não repetir informação equivalente')
        else:
            self.ok('Tabela de resíduo sem coluna "% sem resposta" (regra de não repetição)')

        # regra de 10/07/2026: o relatório não explica ao cliente o próprio processo
        if 'não recebe novos leads' in html:
            self.fail('Texto explicando o funcionamento do fluxo ao cliente '
                      '("não recebe novos leads...") — o cliente conhece o próprio processo')
        else:
            self.ok('Sem explicações do processo do cliente (regra de 10/07/2026)')

        # regra de 06/07/2026: S.02 diária (seg–sex) no semanal; semanal no periodo
        if tipo == 'semanal':
            if 'agrupadas por dia útil' in html:
                self.ok('S.02 em cadência diária seg–sex (padrão do tipo semanal)')
            else:
                self.fail('S.02 não está em cadência diária — o padrão do relatório '
                          'semanal é diário (seg–sex) desde 06/07/2026')
        else:
            if 'agrupadas por semana' in html:
                self.ok('S.02 em cadência semanal (padrão do tipo periodo)')
            else:
                self.fail('S.02 não está em cadência semanal — o padrão do relatório '
                          'de período é semanal')

    def print_report(self, tipo, xlsx_path, graficos_path, html_path=None):
        TIPO_LABEL = {
            'semanal': 'SEMANAL — entrega toda segunda-feira',
            'periodo': 'PERÍODO COMPLETO — entrega primeira segunda do mês',
        }
        print()
        print('=' * 65)
        print(f'  VERIFICAÇÃO DO RELATÓRIO TREBLE — {TIPO_LABEL[tipo]}')
        print('=' * 65)
        print(f'  xlsx:     {xlsx_path}')
        print(f'  graficos: {graficos_path}')
        if html_path:
            print(f'  html:     {html_path}')
        print()

        current = None
        for section, status, msg in self.items:
            if section != current:
                print(f'  [{section}]')
                current = section
            if status == 'ok':
                print(f'    OK     {msg}')
            elif status == 'fail':
                print(f'    FALHA  {msg}')
            else:
                print(f'    AVISO  {msg}')

        print()
        print('=' * 65)
        if self.n_fail == 0:
            status_line = f'  RESULTADO: {self.n_ok} OK | {self.n_warn} avisos'
            print(status_line)
            print('  PRONTO PARA ENTREGA')
        else:
            print(f'  RESULTADO: {self.n_fail} FALHA(S) | {self.n_ok} OK | {self.n_warn} avisos')
            print('  NAO ENTREGAR — resolver falhas primeiro')
            print()
            print('  Falhas encontradas:')
            for _, status, msg in self.items:
                if status == 'fail':
                    print(f'    -> {msg}')
        print('=' * 65)
        print()


def main():
    ap = argparse.ArgumentParser(
        description='Verifica integridade do relatório Treble antes da entrega.'
    )
    ap.add_argument(
        '--tipo', choices=['semanal', 'periodo'], default='semanal',
        help='Tipo do relatório: semanal (7 dias) ou periodo (múltiplos meses)'
    )
    ap.add_argument('--xlsx',     help='Caminho do xlsx de análise (sobrescreve padrão)')
    ap.add_argument('--graficos', help='Pasta de gráficos (sobrescreve padrão)')
    ap.add_argument('--html',     help='Caminho do HTML gerado (opcional)')
    args = ap.parse_args()

    default_xlsx, default_graficos = DEFAULTS[args.tipo]
    xlsx_path     = args.xlsx     or default_xlsx
    graficos_path = args.graficos or default_graficos

    chk = Checker()
    chk.check_xlsx(xlsx_path, args.tipo)
    chk.check_pngs(graficos_path, args.tipo)
    if args.html:
        chk.check_html(args.html, args.tipo)

    chk.print_report(args.tipo, xlsx_path, graficos_path, args.html)
    sys.exit(1 if chk.n_fail > 0 else 0)


if __name__ == '__main__':
    main()
